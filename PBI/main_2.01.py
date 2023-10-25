import pandas as pd
import numpy as np
from openpyxl import load_workbook 
import os
import time

#import collections
#from utils.NAME_COLS import d_name_cols_201910
#import xlwt

""" PATH_HEAD """
#PATH_HEAD = r'C:\Users\tianyunchuan\iCloudDrive\_spyder_\survey\qcToTableau\Nissin'
#PATH_HEAD = r'C:\_python\PowerBI_pyProc\proc_sony_tv'
PATH_HEAD = r'C:\_TianProc\_spyder_\survey\qcToTableau\proc_sony_tv'
if not os.path.exists(r'{}\data_proc'.format(PATH_HEAD)):
    os.mkdir(r'{}\data_proc'.format(PATH_HEAD))

if not os.path.exists(r'{}\data_result_csv'.format(PATH_HEAD)):
    os.mkdir(r'{}\data_result_csv'.format(PATH_HEAD))
    
    
#folder_path = r'{}\data_proc'.format(PATH_HEAD)  # 指定文件夹路径

# 获取文件夹中的所有文件和子文件夹
def removeFile():
    for root, dirs, files in os.walk(r'{}\data_proc'.format(PATH_HEAD)):
        for file in files:
            file_path = os.path.join(root, file)
            try:
                # 删除文件
                os.remove(file_path)
                print(f'已删除文件: {file_path}')
            except OSError as e:
                print(f'删除文件失败: {file_path}, 错误信息: {e}')
removeFile()


""" 全局变量、函数 """
df_trackingMap = pd.read_excel(r'{}\data\{}.xlsx'.format(PATH_HEAD,'trackingMap'), sheet_name='datamap', keep_default_na=False).applymap(str)
l_name_survey = [s for s in list(df_trackingMap.columns) if str(s).startswith('20')]

## Test
#df_trackingMap = df_trackingMap.loc[list(range(0,18)) + list(range(34,37))]
#df_trackingMap = df_trackingMap.loc[list(range(0,18)) ]

#### 2023-1014
df_itemNameReplace = pd.read_excel(r'{}\data\{}.xlsx'.format(PATH_HEAD,'trackingMap'), sheet_name='itemNameReplace', keep_default_na=False).applymap(str)
d_itemReplace = dict(zip(df_itemNameReplace['var_pre'],df_itemNameReplace['var_after']))


## Test
#l_name_survey = l_name_survey[:2]


start = time.perf_counter()
""" 生成 SA, MA 原始数据 """
name_survey = l_name_survey[0]
for name_survey in l_name_survey:
    print(name_survey)
#    name_survey = l_name_survey[1]
#def gen_d_df(name_survey):
    """ 生成词典 设问号：通番 """ 
    df_tmp = df_trackingMap[df_trackingMap[name_survey]!='']
    d_name_cols = dict(zip(df_tmp[name_survey],df_tmp['通番']))
    
    
    """ 获取QC数据 """
    lay = pd.read_excel(r'{}\data\raw_{}.xlsx'.format(PATH_HEAD,name_survey), sheet_name='LayOut', keep_default_na=False).applymap(str)
    #### 2023-1014
    [lay.replace(to_replace=_k, value =_v, inplace=True) for _k, _v in d_itemReplace.items()]

    d_tmp = {}
    [d_tmp.update({s:str(s)}) for s in list(lay.columns)]
    lay.rename(columns=d_tmp, inplace=True) 

    raw = pd.read_excel(r'{}\data\raw_{}.xlsx'.format(PATH_HEAD,name_survey), sheet_name='Raw', keep_default_na=False).applymap(str)
    raw = raw.loc[:,:].replace(r'','*') 
    
    ## 正式样本 !!! ->>>>>>>>>>>>   删除被哟个样本、SC样本
    raw = raw[raw['dcid']!='999']
    
    ## 取对象列
    raw = raw[list(d_name_cols.keys())]
    
    ## 追加Uniqueid列
    raw['Uniqueid'] = [r'{}-{}'.format(name_survey, s) for s in raw['SAMPLEID']]
    
    ## 追加Survey列
    raw['Survey'] = name_survey
    
    ## reset index
    raw.reset_index(drop=True, inplace=True)
    
    ## 生成colname列表
#    l_name_cols = [s for s in d_name_cols.keys() if s not in ['SAMPLEID', 'ANSWERDATE', 'dcid']]
    
    ## 2023-0819加入dcid
    l_name_cols = [s for s in d_name_cols.keys() if s not in ['SAMPLEID', 'ANSWERDATE', ]]
    
    ## 生成字典（选项、设问种类）
    d_qType = dict(zip(lay['Variable'], lay['Answer type']))
    d_qType = dict(zip(l_name_cols, [d_name_cols.get(s).split(' ')[1] for s in l_name_cols]))
    d_qVar = {}
    for _s1 in l_name_cols:
        z_tmp = lay[lay['Variable'] == _s1]
        l_tmp = []
        for _i2 in range(1,251):
            if len(list(z_tmp[str(_i2)])[0])==0:
                break
            l_tmp.append(list(z_tmp[str(_i2)])[0])
    #    l_tmp.insert(0, '*')
        d_qVar.update({_s1:l_tmp})
    
     
    raw.rename(columns=d_name_cols, inplace=True) 
    
    #raw.to_excel(r'{}\data\toTableau_{}.xlsx'.format(PATH_HEAD,name_survey), sheet_name='Raw',index=False)
    
    """ SA -> excel单文件 """
    def gen_tableau_data_sa(data, name_col):
        print('gen_tableau_data_sa--{}'.format(name_col))
#        if d_qType.get(name_col)=='SA':
#        name_col = 'BDALL'
#        data = raw.copy()
        if d_qType.get(name_col)=='SAR':
            for _i1, _s1 in enumerate(data[d_name_cols.get(name_col)]):
                if data[d_name_cols.get(name_col)][_i1] != '*':
                    data[d_name_cols.get(name_col)][_i1]=d_qVar.get(name_col)[int(data[d_name_cols.get(name_col)][_i1])-1]
            data[d_name_cols.get(name_col)] = data[d_name_cols.get(name_col)].replace('*','')
        elif (d_qType.get(name_col)=='MTM') or (d_qType.get(name_col)=='MTS') or (d_qType.get(name_col)=='MAC'):
            del data[d_name_cols.get(name_col)]
#        else:
#            pass
    
    raw_sa = raw.copy()
    [gen_tableau_data_sa(raw_sa, name_col) for name_col in l_name_cols]
#    raw_sa.to_excel(r'{}/data_proc/raw_sa_{}.xlsx'.format(PATH_HEAD,name_survey),index=False)
    raw_sa.to_csv(r'{}/data_proc/raw_sa_{}.csv'.format(PATH_HEAD,name_survey),index=False,encoding="utf_8_sig")
    
    """ MA -> excel多文件 """
    
    def gen_tableau_data_ma(data, name_col):
        print('gen_tableau_data_ma--{}'.format(name_col))
        if (d_qType.get(name_col)=='MTM') or (d_qType.get(name_col)=='MTS') or (d_qType.get(name_col)=='MAC'):
            df_concat = pd.DataFrame()
            for _i1, _s1 in enumerate(raw[d_name_cols.get(name_col)]):
                if _s1!='*':
                    l_1 = _s1.split(',')
                    l_1 = [s for s in l_1 if s!='']
                    l_1 = list((set(l_1)))
                    l_1 = [d_qVar.get(name_col)[int(s)-1] for s in l_1]
                    
                    df_tmp = pd.DataFrame({d_name_cols.get(name_col):l_1,'Uniqueid':[raw['Uniqueid'][_i1] for i in range(0,len(l_1))]})
                    df_concat = pd.concat([df_concat,df_tmp])
#            df_concat.to_excel(r'{}/data_proc/raw_ma_{}_{}.xlsx'.format(PATH_HEAD,name_survey,d_name_cols.get(name_col)),index=False)
            df_concat.to_csv(r'{}/data_proc/raw_ma_{}_{}.csv'.format(PATH_HEAD,name_survey,d_name_cols.get(name_col)),index=False,encoding="utf_8_sig")
    ##  def exe !
    [gen_tableau_data_ma(raw, name_col) for name_col in l_name_cols]
               
delta = time.perf_counter() - start
print("程序运行的时间是：{}秒".format(delta))     





    
    
    
    
    





    

    
""" to result file """    
    
start = time.perf_counter()   
""" 建立result excel空表 """
if not os.path.exists(r'{}\data_proc'.format(PATH_HEAD)):
    os.mkdir(r'{}\data_proc'.format(PATH_HEAD))
df_tmp = pd.DataFrame()
df_tmp.to_excel(r'{}\data\result_toTableau.xlsx'.format(PATH_HEAD))

book = load_workbook(r'{}\data\result_toTableau.xlsx'.format(PATH_HEAD))
writer = pd.ExcelWriter(r'{}\data\result_toTableau.xlsx'.format(PATH_HEAD), engine='openpyxl')
writer.book = book



 
    

""" 预处理文件列表 """
l_file_all = os.listdir(r'{}\data_proc'.format(PATH_HEAD))
l_df_all = []


""" 生成 Result 文件 1. raw_sa """
l_file = [s for s in l_file_all if 'raw_sa' in s]
l_df = [pd.read_csv(r'{}\data_proc\{}'.format(PATH_HEAD,s)) for s in l_file]
raw_tmp = pd.concat(l_df,axis=0,ignore_index=True, sort=False)   
l_df_all.append(raw_tmp) 

l_tmp = [s.replace('SAR 縦積なし ','') for s in raw_tmp.columns]
l_tmp = [s.replace('Q00','Q') for s in l_tmp]
d_tmp = dict(zip(list(raw_tmp.columns),l_tmp))

raw_tmp.rename(columns=d_tmp, inplace=True)  

raw_tmp.to_excel(writer,sheet_name='_raw_sa',index=False)
writer.save()


#### Test 1014
#raw_tmp.to_csv(r'{}\data_result_csv/raw_sa.csv'.format(PATH_HEAD),index=False,encoding="utf_8_sig")





""" 生成 Result 文件 2. ma多选题 (ok, 无需縦積み) """
l_file = [s for s in l_file_all if 'MAC' in s]
l_map_tmp = [s for s in list(set(df_trackingMap["通番"])) if 'MAC' in s]
l_map_qContent = [s.split(" ")[3] for s in l_map_tmp]
for _s1 in l_map_qContent:
    print(_s1)
    l_file_tmp = [s for s in l_file if _s1 in s]
    l_df = [pd.read_csv(r'{}\data_proc\{}'.format(PATH_HEAD,s)) for s in l_file_tmp]
    raw_tmp = pd.concat(l_df,axis=0,ignore_index=True, sort=False)    
    l_df_all.append(raw_tmp) 
    raw_tmp.to_excel(writer,sheet_name=_s1,index=False)
    writer.save()
#    raw_tmp.to_csv(r'{}\data_result_csv/{}.csv'.format(PATH_HEAD,_s1),index=False,encoding="utf_8_sig")



""" 生成 Result 文件 3. 需縦積み """
l_file = [s for s in l_file_all if '縦積MT' in s]
l_map_qContent = [s for s in list(set(df_trackingMap["補助3"])) if '縦積MT' in s]
#l_map_qContent = [s.split(" ")[3] for s in l_map_tmp]
#_s1 = l_map_qContent[1]
for _s1 in l_map_qContent:
    try:
        print(_s1)
        l_file_tmp = [s for s in l_file if _s1 in s]
        l_df = [pd.read_csv(r'{}\data_proc\{}'.format(PATH_HEAD,s)) for s in l_file_tmp]
        for df in l_df:        
            df['Label'] = list(df.columns)[0].split(' ')[-1]
            df.rename(columns={list(df.columns)[0]: _s1.replace('縦積MT','')}, inplace=True) 
        raw_tmp = pd.concat(l_df,axis=0,ignore_index=True, sort=False)    
        l_df_all.append(raw_tmp) 
        raw_tmp.to_excel(writer,sheet_name=_s1.replace('縦積MT',''),index=False)
        writer.save()
#        raw_tmp.to_csv(r'{}\data_result_csv/{}.csv'.format(PATH_HEAD,_s1),index=False,encoding="utf_8_sig")

        #### 2023-1014
        """ 拆解KPI """
        if _s1 == '縦積MTKPI':

            s_kpi_all = r'认知,考虑,首选,不考虑,最终购买,之前使用,目前使用'
            s_kpi_1 = r'认知,考虑,首选,不考虑'
            s_kpi_2 = r'最终购买,之前使用,目前使用'            
            l_kpi_all = s_kpi_all.split(",")
            
            for _k1 in l_kpi_all:
                print(_k1)
                
                raw_tmp_1 = raw_tmp[raw_tmp['Label']==_k1]
                raw_tmp_1.to_excel(writer,sheet_name='KPI_{}'.format(_k1),index=False)
                writer.save()
            raw_tmp_1 = raw_tmp[raw_tmp['Label'].isin(s_kpi_1.split(','))]
            raw_tmp_1.to_excel(writer,sheet_name='KPI_step1',index=False)
            writer.save()
            print('KPI_step1')
        
            raw_tmp_1 = raw_tmp[raw_tmp['Label'].isin(s_kpi_2.split(','))]
            raw_tmp_1.to_excel(writer,sheet_name='KPI_step2',index=False)
            writer.save()
            print('KPI_step2')

    except:
        print('{}->error'.format(_s1))

delta = time.perf_counter() - start
print("程序运行的时间是：{}秒".format(delta))     
writer.save()



""" 删除 Sheet1 """
book = load_workbook(r'{}\data\result_toTableau.xlsx'.format(PATH_HEAD))
del book['Sheet1']
writer.book = book
writer.save()


""" Excel加入Sheet_list工作表 """
excel_file = pd.ExcelFile(r'{}\data\result_toTableau.xlsx'.format(PATH_HEAD))
sheet_names = excel_file.sheet_names
df_sheet_list = pd.DataFrame({'编号': range(1, len(sheet_names) + 1), 'sheet_list': sheet_names})
df_sheet_list.to_excel(writer,sheet_name="_navigation" ,index=False)
writer.save()








