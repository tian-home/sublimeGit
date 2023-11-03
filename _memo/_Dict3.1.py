#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Dec  6 08:07:38 2019

@author: yunchuantian
"""

from pandas import DataFrame 
import pandas as pd
import numpy as np


""" python doc """
## 生成doc   cmd 状态!!!
# C:\_TianProc\_spyder_\_test\bili_class>python -m pydoc -p 1234 -> Server ready at http://localhost:1234/
# 不需要指定端口  (base) C:\_TianProc\_spyder_\_test\bili_class>python -m pydoc -b
# python -m pydoc -w ch037    # 当前路径生成文档
# python --help



""" [sample数据取样] """
import _utils_.DB_MySql.Get_Data_From_MySql as getSQL
getSql = getSQL.Get_Data_From_MySql('world', 'country')
df = getSql.go()

DataFrame.sample(n=None, frac=None, replace=False, weights=None, random_state=None,axis=None)

# n： 要抽取的行数,需为整数值
# frac：抽取的比列,需为小数值，比方说我们想随机抽取30%的数据，则设置frac=0.3即可。
# replace：抽样后的数据是否代替原DataFrame()，默认为False
# weights：默认为等概率加权
# random_state：随机种子，本质是一个控制器，设置此值为任意实数，则每次随机的结果是一样的
# axis：抽取数据的行还是列，axis=0的时是抽取行，axis=1时是抽取列

df1 = df.sample(n=10,random_state=123,axis=0)
df2 = df.sample(frac=0.01,axis=0)


""" 判断函数 isinstance """
# 对象的三个特征：
# value
# id
# type
'' is None
np.nan is None
None is None
isinstance(object, classinfo)

a = 2
isinstance (a,int)
True
isinstance (a,str)
False
isinstance (a,(str,int,list))    # 是元组中的一个返回 True
True
isinstance ('',(str,int)) # 是否是其中的一个


# type() 与 isinstance()区别：
class A:
    pass
 
class B(A):
    pass
 
isinstance(A(), A)    # returns True
type(A()) == A        # returns True
isinstance(B(), A)    # returns True
type(B()) == A        # returns False

# int，float，bool，complex，str(字符串)，list，dict(字典)，set，tuple


""" None """
#   None 空  ≠  空字符串 空的列表、0、False
#   以上类型、取值都不相等
#   None的type是 ‘NoneType’
a = ''
b = False
c = []

a==None  #False
b==None  #False
c==None  #False
a is None
b is None
c is None

type(None)  #NoneType

#   判空操作 建议 if a:   if not a (None,'',[],False都被包含)
if a:
    print('有值')

if not a:               # (None,'',[],False都被包含)
    print('无')

# 正无穷赋值
best_entropy = float('inf')


""" [os] """
# 获取文件夹中所有文件的文件名


path = r'C:\Users\tianyunchuan\iCloudDrive\Desktop\_win_Trans_mac_\_trackingDataConcat\data_test'
list_dir = os.listdir(path)


fileNames_txt =  [s for s in os.listdir(path) if s.split('.')[-1] == 'txt']

## 如果不存在文件夹，则创建文件夹
if not os.path.exists('model'):
    os.mkdir('model')

## cpu核数
os.cpu_count()

# 当前文件路径
import os, sys
print(os.path.dirname(__file__))
# sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.path.abspath(__file__)
PATH_HEAD = os.path.abspath("..")

""" [break], [continue] """
for i in range(1,4):
    for letter in 'Python':     # 第一个实例
       if letter == 'h':
          break
       print('当前字母 :', letter)    
    print('-'*20)   

for letter in 'Python':     # 第一个实例
   if letter == 'h':
      continue
   print('当前字母 :', letter)
 
var = 10                    # 第二个实例
while var > 0:              
   var = var -1
   if var == 5:
      continue
   print('当前变量值 :', var)
print("Good bye!")




## 只读取文件夹
root =  r'C:\Users\tianyunchuan\iCloudDrive\Desktop\_jupyter_\ML_PyTorch_class_163_Singapore'
name2label = {}
for name in sorted(os.listdir(os.path.join(root))):
    if not os.path.isdir(os.path.join(root, name)):
        continue
    name2label[name] = len(name2label.keys())
print(name2label)

## glob 提取文件名包含'*'函数
import glob
list = glob.glob('*Py*')

# 1、通配符 星号(*)匹配零个或多个字符
import glob
for name in glob.glob('_utils_/*'):
    print (name)

# 列出子目录中的文件，必须在模式中包括子目录名：
  #用子目录查询文件
print ('Named explicitly:')
for name in glob.glob('dir/subdir/*'):
    print ('\t', name)
#用通配符* 代替子目录名
print ('Named with wildcard:')
for name in glob.glob('dir/*/*'):
    print ('\t', name)  

# 2、单个字符通配符  用问号(?)匹配任何单个的字符。
import glob
for name in glob.glob('dir/file?.txt'):
    print (name)   
    
# 3、字符范围 当需要匹配一个特定的字符，可以使用一个范围
for name in glob.glob('dir/*[0-9].*'):
    print (name)

## os.path.join
import os

Path1 = 'home'
Path2 = 'develop'
Path3 = 'code'

Path10 = Path1 + Path2 + Path3
Path20 = os.path.join(Path1,Path2,Path3)
print ('Path10 = ',Path10)
print ('Path20 = ',Path20)

# result
# Path10 = homedevelopcode
# Path20 = home\develop\code


""" [逻辑判断] """ 
#    用or逻辑判断 条件赋值None为空 !!!!!!
pages = None
pages = None
pages = 999
book = {'pages': pages or ''} 
book['pages']






""" [time] """
import time, datetime
import pandas as pd
from datetime import timedelta

time.ctime()    # 'Wed Dec 19 22:05:05 2018'

df['a']=pd.to_datetime(df['a'])

t1=datetime.datetime.strptime(df_0.投票时间[0],'%Y-%m-%d %H:%M:%S')
t2=datetime.datetime.strptime(df_0.投票时间[1],'%Y-%m-%d %H:%M:%S')
t1-t2

#时间计算1 不推荐
start = time.clock()
df['人均']=df.人均.str.strip('￥')    
end = time.clock()
str(end-start)
time.sleep(20)

#时间计算2 推荐  计时
import time
limit = 10*100*1000
start = time.perf_counter()
 
while True:
    limit -= 1
    if limit <= 0:
        break
delta = time.perf_counter() - start
print("程序运行的时间是：{}秒".format(delta))


datetime.MAXYEAR    #返回9999
datetime.MINYEAR    #返回1
td=datetime.date.today()
td.year
td.month
td.day
td.weekday()
td.isoweekday()
birth=datetime.date(1972,10,20)

t=datetime.datetime.now()
t.hour
t.minute
t.second


# 过程时间计算, 计算时间
from datetime import datetime
#import datetime
start_time = datetime.now()
[i for i in range(1,10000000)]
print("循环耗时：{}".format((datetime.now() - start_time).total_seconds()))


#格式转换
#strptime 字符串→日期，strftime 日期→字符串
s='1972-10-20'
ti=datetime.datetime.strptime(s,'%Y-%m-%d')     #注：年Y要大写！！！！
s='1972/10/20'
ti=datetime.datetime.strptime(s,'%Y/%m/%d')  

datetime.datetime(2017,7,18,9,28,56,000000)
now=datetime.datetime.now()
now.strftime('%Y---%m---%d')

d0=datetime.datetime(2018,3,5,10,30)
d1=datetime.datetime(2017,7,18,9,20)
diff=d0-d1
a=diff.days
a1=diff.seconds
a2=diff.total_seconds()

o=datetime.datetime(2008,8,8,20,8)
o+datetime.timedelta(days=100)
result=o+datetime.timedelta(days=100)
result1=o+datetime.timedelta(days=-100)
result2=o+datetime.timedelta(seconds=-10000)



# 设置时间跨度 timedelta
from datetime import timedelta
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)

# 时间分割区域
##pip install datestuff==0.2.0
from datestuff import DateRange
from datetime import date, timedelta
dr = DateRange(start=date(2016, 1, 1),
               stop=date(2016, 4, 30),
               step=timedelta(days=7))
[[dr[i], dr[i+1]-timedelta(days=1)] for i in range(0, len(dr)-1)]
d1 = dr[0]

## 时间分割2.
dr = pd.date_range('2021-6-30 12:00:00',periods=365*5,freq='-12H') 

import pandas as pd
dates = pd.date_range(start="2018-10-01", end="2018-10-12")
dates1 = pd.date_range(start="2018-10-01", periods=6)
print(dates)
print(dates1)
DatetimeIndex(['2018-10-01', '2018-10-02', '2018-10-03', '2018-10-04',
               '2018-10-05', '2018-10-06', '2018-10-07', '2018-10-08',
               '2018-10-09', '2018-10-10', '2018-10-11', '2018-10-12'],
              dtype='datetime64[ns]', freq='D')
DatetimeIndex(['2018-10-01', '2018-10-02', '2018-10-03', '2018-10-04',
               '2018-10-05', '2018-10-06'],
              dtype='datetime64[ns]', freq='D')


import pandas as pd
dates = pd.date_range(start="2018-10-01", periods=8, freq="B")
dates1 = pd.date_range(start="2018-10-01", periods=8, freq="H")
dates2 = pd.date_range(start="2018-10-01", periods=8, freq="S")



import pandas as pd
dates = pd.date_range(start="2018-10-01 03:00:00", periods=8, freq="H", normalize=True)
dates1 = pd.date_range(start="2018-10-01 03:00:00", periods=8, freq="H", normalize=False)

pd.date_range('20100101', '20101201', freq='BM')

## https://blog.csdn.net/weixin_42102634/article/details/113479508

import time
 
struct_time = time.strptime("30 Nov 00", "%d %b %y")
#Out[8]: time.struct_time(tm_year=2000, tm_mon=11, tm_mday=30, tm_hour=0, tm_min=0, tm_sec=0, tm_wday=3, tm_yday=335, tm_isdst=-1)
time.strptime("30 11 00", "%d %m %y")
time.strptime("21 02 05", "%y %m %d")










""" [str] str补全， 字符串补全 """
#ljust（len，str）字符向左对齐，用str补齐长度
#rjust（len，str）字符向右对齐，用str补齐长度
#rjust（len，str）字符中间对齐，用str补齐长度


'bbb'.ljust(10,'a')
# 输出：bbbaaaaaaa
'bbb'.rjust(10,'a')
# 输出：aaaaaaabbb
'bbb'.center(10,'a')
# 输出：aaabbbaaaa

# zfill（width）指定字符串长度，右对齐，前面补充0
z = '2'.zfill(5)
# 输出：00002

####        /t/n/r strip
#strip 同时去掉左右两边的空格
#lstrip 去掉左边的空格
#rstrip 去掉右边的空格

#具体示例如下：
a=" gho stwwl "
a.lstrip() 'gho stwwl '
a.rstrip() ' gho stwwl'
a.strip() 'gho stwwl'

#声明：s为字符串，rm为要删除的字符序列
#s.strip(rm) 删除s字符串中开头、结尾处，位于 rm删除序列的字符
#s.lstrip(rm) 删除s字符串中开头处，位于 rm删除序列的字符
#s.rstrip(rm) 删除s字符串中结尾处，位于 rm删除序列的字符

#注意：1. 当rm为空时，默认删除空白符（包括'\n', '\r', '\t', ' ')

a = ' 123'
a.strip()
'123'
>>> a='\t\tabc'
'abc'
a = 'sdff\r\n\t'
a.strip('\n')
a.strip('\t')

## 字符包含any
p = "Tom is a boy,Lucy is a girl,they all like english!"
keywords= 'Tom,Lucy'
excludes = ['english','math']
any([w in p and w for w in keywords.split(',')])
any(e in p for e in excludes)

## 字符包含all
p = "Tom is a boy,Lucy is a girl,they all like english!"
keywords= 'Tom,Lucy'
filters= ["boy","like"]
all(f in p for f in filters)
all([w in p and w for w in keywords.split(',')])


""" [list] """
## 列表去除空白
s = '我们不再回到过去，  嘿'
[x for x in s]
[x for x in s if x!=' ']


np.random.normal(3000, 25.1056, 90)  # 均值3000，方差，样本量

#### 删除元素
lang = ["Python", "C++", "Java", "PHP", "Ruby", "MATLAB"]
## del：根据索引值删除元素
# del 可以删除列表中的单个元素，格式为：
del listname[index]
del listname[start : end]

del lang[2]
del lang[-2]
del lang[1: 4]

## pop()：根据索引值删除元素
nums = [40, 36, 89, 2, 36, 100, 7]
nums.pop(3)
nums.pop()

## remove()：根据元素值进行删除
nums = [40, 36, 89, 2, 36, 100, 7]
nums.remove(36)
nums.remove(36)
nums.remove(78)

## clear()：删除列表所有元素
nums.clear()

# list均值
import numpy as np
a = [2,4,6,7,10]
average_a = np.mean(a)
median_a = np.median(a)

S=[55,66,45,32,67,89,99,67,8]
filter(lambda s: s>70, S)
list(filter(lambda s: s>70, S))


#在列表末尾添加新的对象
list.append(obj)
# 统计某个元素在列表中出现的次数
list.count(obj)
# 在列表末尾一次性追加另一个序列中的多个值（用新列表扩展原来的列表）
list.extend(seq)
# 从列表中找出某个值第一个匹配项的索引位置
list.index(obj)
# 将对象插入列表
list.insert(index, obj)
# 移除列表中的一个元素（默认最后一个元素），并且返回该元素的值
list.pop([index=-1])
# 移除列表中某个值的第一个匹配项
list.remove(obj)
# 反向列表中元素
list.reverse()
# 对原列表进行排序
list.sort(cmp=None, key=None, reverse=False)

#### 简繁转换 简体, 繁体
# https://zhuanlan.zhihu.com/p/55973055
# zh-tw 台灣正體
# zh-hk 香港繁體
# zh-sg 马新简体
# zh-hans 简体
# zh-hant 繁體

from zhconv import convert
convert('Python是一种动态的、面向对象的脚本语言', 'zh-hant')
'Python是一種動態的、面向對象的腳本語言'

convert('Python是一種動態的、面向對象的腳本語言', 'zh-hans')
convert('尚多拉美髮サロン', 'zh-hans')
convert('BARBERKING男士理髪馆', 'zh-hans')

# list插入
# append多用于把元素作为一个整体插入
# insert多用于固定位置插入
# extend多用于list中多项分别插入


## 分区间
a = [30,'',45]
pd.cut(a,bins=[0,10,20,40,60,100],labels=["婴儿","青年","中年","壮年","老年"])

import pandas as pd 
import numpy as np 
  
  
df= pd.DataFrame({'number':np.random.randint(1, 100, 10)}) 
df['bins'] = pd.cut(x=df['number'], bins=[1, 20, 40, 60,  
                                          80, 100]) 
print(df) 
  
# We can check the frequency of each bin 
print(df['bins'].unique())


"""  [dict] """
#    字典序列化递归 ???
import json
json.dumps(book, default=lambda  o:o.__dict__)

## 列表转字典
list1 = ['key1','key2','key3']
list2 = ['1','2','3']
d = dict(zip(list1,list2))


## 修改key
dict={'a':1, 'b':2}
dict["c"] = dict.pop("a")

## dict排序
dd = {'banana': 3, 'apple':4, 'pear': 1, 'orange': 2}
#按key排序
kd = dict(collections.OrderedDict(sorted(dd.items(), key=lambda t: t[0])))
#按照value排序
vd = collections.OrderedDict(sorted(dd.items(),key=lambda t:t[1]))

## dict列表排序
# -*- coding:utf-8 -*-
anchors = [{'title': '蒜香莲藕虾', 'mat': '用料：莲藕、虾、蒜、葱、酱油、盐、白糖、水、植物油', 'collect': 28}, {'title': '小炒金钱肚', 'mat': '用料：金钱肚、青椒、胡萝卜、洋葱、小米辣、生姜、大蒜头、油、生抽、白糖、盐、黄酒', 'collect': 98}, {'title': '手抓牛肉', 'mat': '用料：牛肉、胡萝卜、豆角、香菇、蒜、姜、蒜苗', 'collect': 19}, {'title': '茯苓姬松茸排骨汤', 'mat': '用料：姬松茸、莲子、百合、茯苓、生姜、排骨、盐、纯净水、料酒', 'collect': 92}, {'title': '凉拌娃娃菜', 'mat': '用料：娃娃菜、小米椒、蒜、醋、生抽、香油', 'collect': 95}]

# anchors = sorted(anchors,key=collect)
anchors.sort(key=lambda x:x["collect"], reverse=True)

for s in anchors:
    print(s)



## 字典合并
    
# 两个字典
dict1 = {'a': 10, 'b': 8} 
dict2 = {'d': 6, 'c': 4} 
dict2.update(dict1)

# 优雅方式
result_dict = {**dict_1, **dict_2}


## 字典追加为列表
d = {
        'nickname':nickname,
        'review_content':review_content
        }
info.append(d)

## dict映射替代switch case
day=1
def get_sunday():
    return 'sunday'

def get_monday():
    return 'monday'

def get_tuesday():
    return 'tuesday'

def get_default():
    return 'unknow'

switcher={
        0:get_sunday,
        1:get_monday,
        2:get_tuesday
        }

day_name=switcher.get(day,get_default)()
print(day_name)

## 字典的列表推导
characters={'马超':99, '许褚':98, '赵云':100, '张郃':90}
[key for key,value in characters.items()]
{value:key for key,value in characters.items()}
tuple(key for key,value in characters.items())


## 增加一个或多个元素
d = {'a': 1}
d.update(b=2)  #也可以 d.update({‘b’: 2})
d.update(c=3, d=4)

d['e'] = 5
d.update({'f': 6, 'g': 7})  #即d.update(字典)
## 删除一个或多个元素
x = {1: 2, 3: 4, 4: 3, 2: 1, 0: 0}
x.pop(1)   # pop(key)
del x[3]


## 调用
d['key_name']

## KeyError处理方法
user = dict(name="brainliao", age=32)
print(user.get("sex", "男"))

[dict_columns.get(s, s) for s in list_raw_columns]

## defaultdict
from collections import defaultdict

dict1 = defaultdict(int)
dict2 = defaultdict(set)
dict3 = defaultdict(str)
dict4 = defaultdict(list)
dict1[2] ='two'
print(dict1[1])
print(dict2[1])
print(dict3[1])
print(dict4[1])

## dictTODataFrame  http://www.6quant.com/?p=2391

dict_data = {u'2012-06-08': 388,
 u'2012-06-09': 388,
 u'2012-06-10': 388,
 u'2012-06-11': 389,
 u'2012-06-12': 389,
 u'2012-06-13': 389}

## keys转为pandas列名，values转为values
pd.DataFrame.from_dict(dict_data,orient='index').T

## 转变为Series
s  = pd.Series(dict_data,index=dict_data.keys())

## 转变为DataFrame
s = pd.DataFrame(list(dict_data.items()), columns=['Date', 'DateValue'])

s = pd.DataFrame.from_dict(dict_data, orient='index')

s = pd.DataFrame(dict_data.items(), columns=['Date', 'DateValue'])
s['Date'] = pd.to_datetime(s['Date'])


## df转dict
import pandas as pd
df = pd.DataFrame({'name':['赵云','马超','许褚'],"class":[11,22,33],"price":[111,222,333]})
z = df.to_dict(orient="list")
z = df.to_dict(orient="series")
z = df.to_dict(orient="split")
z = df.to_dict(orient="records")
z = df.to_dict(orient="index")


""" [lambda] 表达式 """
#   基本形式
action = lambda a, b: a + b
action(100,300)

#   lambda形成列表
a,b=100,3
func_list=[lambda x,y:x+y,lambda x,y:x-y,lambda x,y:x*y,lambda x,y:x/y]
for func in func_list:
    print(func(a,b))



""" [三元表达式] """
import numpy as np
999 if 6>3 else 111
arr = np.random.randint(1, 100, size=100)
arr_booleans = np.where(arr>30, 2, -2)




""" [推导[] [映射] [map] [过滤] [filter] """
## 推导
lst = ['a','b','x','d','aa','abc','z']
[f'first{s}' for s in lst]
#   条件推导
[f'first{s}' for s in lst if 'a' in s]
[x*x for x in np.arange(10) if x%3==2]

## 映射 map
s=[55,66,45,32,67,89,99,67,8]
map(lambda s: s**3, s)
list(map(lambda s: s**3, s))

df['gender_cate'].map({'男': 1, '女': 2}).astype(int)

stu=['Jerry','Tom','Mark']
list(map(lambda s: s.upper(), stu))

mapper = {'Fully Paid':0, 'Charged off':1}
df['indicator'] = df['loan_status'].map(mapper)

df.replace(['10+ years', '4 years', '<1 years'], [10, 4, 1])

## 过滤 filter
S=[55,66,45,32,67,89,99,67,8]
filter(lambda s: s>70, S)
list(filter(lambda s: s>70, S))

stu=['Jerry','Tom','Mark']
list(filter(lambda s: 'e' in s, stu))

## 列表推导&filter
[x>60 for x in s]
list(filter(lambda x: x>=70,s))
[x for x in s if x>70]

## 去重列表推导&filter
s.append(66)
{x for x in s}
stu.append('Jerry')
{x for x in stu}

## 推导+if
tag_list=['职场', ' 1 评论 ', '程序员', '职场']
[element for element in tag_list if not element.strip().endswith('评论')]

## apply
df = DataFrame(np.random.randn(4, 3), columns=list('bde'), index=['Utah', 'Ohio', 'Texas', 'Oregon'])
f = lambda x: x.max() - x.min()
df.apply(f)
df.apply(f, axis=0)
df.apply(f, axis=1)

df = pd.DataFrame([[4, 9]] * 3, columns=['A', 'B'])
df.apply(np.sqrt)  # 作用于所有单元

##　applymap
#如果想让方程作用于DataFrame中的每一个元素，可以使用applymap().用法如下所示
format = lambda x: '%.2f' % x
df.applymap(format)
df.applymap(lambda x: x+1000)

df = pd.DataFrame([[4, 9],] * 3, columns=['A', 'B'])
df1 = df.applymap(str)


## to_replace  https://vimsky.com/examples/usage/python-pandas-dataframe-replace.html
# 我们将在“ df” DataFrame 中将团队“Boston Celtics”替换为“Omega Warrior”
# this will replace "Boston Celtics" with "Omega Warrior" 
df.replace(to_replace ="Boston Celtics", value ="Omega Warrior")
# 范例2：一次替换多个值。使用python列表作为参数
# this will replace "Boston Celtics" and "Texas" with "Omega Warrior" 
df.replace(to_replace =["Boston Celtics", "Texas"],  value ="Omega Warrior")
# 范例3：用-99999值替换 DataFrame 中的Nan值。
df.replace(to_replace = np.nan, value =-99999)

""" [格式、 format] """

## f-string
# 简单使用  f-string用大括号 {} 表示被替换字段，其中直接填入替换内容：
name = 'Eric'
f'Hello, my name is {name}'
number = 7
f'My lucky number is {number}'
# 表达式求值与函数调用 f-string的大括号 {} 可以填入表达式或调用函数，Python会求出其结果并填入返回的字符串内：
f'A total number of {24 * 8 + 4}'
f'Complex number {(2 + 2j) / (2 - 3j)}'

name = 'ERIC'
f'My name is {name.lower()}'

import math
f'The answer is {math.log(math.pi)}'

# 引号、大括号与反斜杠 f-string大括号内所用的引号不能和大括号外的引号定界符冲突，可根据情况灵活切换 ' 和 "：
f'I am {"Eric"}'

# 若 ' 和 " 不足以满足要求，还可以使用 ''' 和 """：
f"He said {"I'm Eric"}"  错误！
f'He said {"I'm Eric"}'  错误！
f"""He said {"I'm Eric"}"""

# 大括号外的引号还可以使用 \ 转义，但大括号内不能使用 \ 转义：
f'''He\'ll say {"I'm Eric"}'''

# f-string大括号外如果需要显示大括号，则应输入连续两个大括号 {{ 和 }}：
f'5 {"{stars}"}'
f'{{5}} {"stars"}'
f
# 上面提到，f-string大括号内不能使用 \ 转义，事实上不仅如此，f-string大括号内根本就不允许出现 \。如果确实需要 \，则应首先将包含 \ 的内容用一个变量表示，再在f-string大括号内填入变量名：
newline = ord('\n')
f'newline: {newline}'

# 多行f-string f-string还可用于多行字符串：
name = 'Eric'
age = 27
f"Hello!" \
... f"I'm {name}." \
... f"I'm {age}."
"Hello!I'm Eric.I'm 27."

f"""Hello!
... I'm {name}.
... I'm {age}."""
"Hello!\n I'm Eric.\n I'm 27."








## 数值格式 千位分隔符,千分位，数值格式, f{}
'{:,}'.format(1000000)
a=5200000000000
b=1453453423.12132424
c=-1232423423.3242342

'数值={:.2f}'.format(a/b)
'数值={:.2%}'.format(a/b)

'{:.2f}'.format(3.1415)
'{:.4f}'.format(3.1)

f'数值={c}'
f'数值={a:f}'
f'数值={b:,f}'
f'数值={b:,.2f}'

f'{c:,.0f}'
f'{c:.2%}'

## 格式_小数点
# https://www.cnblogs.com/Raymon-Geng/p/5784290.html
a,b = 5.026, 5.000
round(a,2)
round(b,3)
'%.2f' % a
'%.2f' % b
float('%.2f' % a)    #type(float('%.2f' % a))
float('%.2f' % b)
from decimal import Decimal
Decimal('5.026').quantize(Decimal('0.00'))
Decimal('5.000').quantize(Decimal('0.00'))



## %数字格式
'%s %s %s' % (1, 2.3, ['one', 'two', 'three'])
'%20.2f' % 1.235        #20个字符，其中小数2位
'%020.2f' % 1.235       #20个字符，其中小数2位，不足用0补
'%(name)s:%(score)06.1f'%{'score':9.5, 'name':'newsim'}  

## 百分比
# https://www.cnblogs.com/nulige/p/6115793.html
tpl = 'percent %.2f %%' % 99.976234444444444444
tpl1 = '%.2f %%' % 99.976234444444444444

'percent: {:.2%}'.format(42/50)

## 填充01-13
[r"Peashooter_{:02d}.png".format(i)  for i in range(0, 13)]

## 小数点显示
import math
math.trunc(b)       #math截断
math.floor(b)       #往下舍弃
math.ceil(b)       #向上取
round(b,2)      #四舍五入 

print('常量 PI 的值近似为： {}。'.format(math.pi))
print('常量 PI 的值近似为： {!r}。'.format(math.pi))
print('常量 PI 的值近似为 {0:.3f}。'.format(math.pi))
print("{:.2f}".format(3.1415926));

print("%6.3f" % 2.3)        #2.300
    # 第一个 % 后面的内容为显示的格式说明，6 为显示宽度，3 为小数点位数，f 为浮点数类型 
    # 第二个 % 后面为显示的内容来源，输出结果右对齐，2.300 长度为 5，故前面有一空格 
print("%+10x" % 10)
    # x 为表示 16 进制，显示宽度为 10，前面有 8 个空格。
print("%-5x" % -10)    
    # %-5x 负号为左对齐，显示宽度为 5，故 -a 后面有 3 个空格

pi=3.1415  
print ("pi的值是%s"%pi)    
print ("pi的值是%.8f"%pi)
    # 上面的 width, precision 为两个整数。我们可以利用 *，来动态代入这两个量。比如：
print("%10.*f" % (4, 1.2))  



## 通过下标
#http://www.cnblogs.com/kaituorensheng/p/5709970.html
'{0[0]} is {0[1]} years old!'.format(['jihite', 4])
'{0} is {1} years old!'.format('jihite', 4)

## 通过关键字
'{name}:{age}'.format(age=4,name='jihite')
'{name}:{age}'.format(age=4,name='jihite',locate='Beijing')

## 通过位置
'{0} is {1}'.format('jihite', '4 years old')
'{0} is {1} {0}'.format('jihite', '4 years old')

## astype  改变类型
df1['price'] = df['price'].astype(str)
df1['price'] = df['price'].astype('int')

## if_else + 迭代[] + float
df['月成交1']=[(float(s.replace('万', ''))*10000 if '万' in s else s) for s in df.月成交]

## 通过Python3.6版本 f
def r(a,b,c):
    print(f'我是{a}名字{b}哈哈s{c}')
r('田','蕴','川')














""" [进制转化] """
'{:b}'.format(10)   # Binary, 二进制
'{:o}'.format(10)   # Octet, 八进制
'{:d}'.format(10)   # Decimal, 十进制
'{:x}'.format(10)   # Hex, 十六进制

bin(10)     #向二进制转化
int(0o7)    #int向十进制转化
hex(0x7)    #向十六进制转化
oct(0b111)  #向八进制转化

#### 16进制 十六进制 提取字符 ！！！！！！！ !!!!!
sixteen_key = '\ue987'
parsed_code = []
for x in sixteen_key:
    parsed_code.append(hex(ord(x))[2:])
parsed_code[0]



""" [字符集， 编码] """
## encode
s='abc'
su=u'abc'
s.encode('utf8')
su.encode('utf8')

s='我用python'
su=u'我用python'
s.encode('utf8')
su.encode('utf8')
s.encode('gbk')

## decode
s='我用python'
su=u'我用python'
s_utf = s.encode('utf8')
s_utf.decode()




""" 数学 math 库 """
import  math
math.exp(100)   #:e的x次幂
math.log(999)     #:对数函数，底数是e
math.log10(567)   #:对数函数，底数是10




""" [内置函数] """
####    1. enumerate （数组、元祖自动生成 顺序index）迭代
[print(index,item) for index,item in enumerate(['a','b','c'])]
[print(index,item) for index,item in enumerate(('a','b','c'))]
str01 = 'abcdeffggg'
for ind,s in enumerate(str01,100):   # 第二参数是起始位置
    print(ind,s)

seq = ['one', 'two', 'three']
for i, element in enumerate(seq):
    print(i, seq[i])

####    2. eval (将文本字符转 转为 算式)
x, y = 5,99
eval('x+y')
eval('x*2+y')

x = [1,2,3,5]
eval('list(map(lambda s :s*s,x))')

####    3. 余数
[i%3 for i in range(1,11)]

[x*x for x in np.arange(10)]
[x*x for x in np.arange(10) if x%3==2]

####    4. set （交集、并集） , type: set
x = set("jihite")
y = set(['d', 'i', 'm', 'i', 't', 'e'])
x & y       #交集
x|y         #并集
x-y         #差
y-x         #差
x^y         #对称差：x和y的交集减去并集

{1,2,3,4,5,} - {2,4,9}      #差集
{1,2,3,4,5,} & {2,4,9}      #交集
{1,2,3,4,5,} | {2,4,9}      #并集
set()                       #定义空集

a=[2,3,4,5,5]
b=[2,5,8]
[val for val in a if val in b]



####    5. strip()函数
# =================
# 声明：s为字符串，rm为要删除的字符序列
# s.strip(rm)        删除s字符串中开头、结尾处，位于 rm删除序列的字符
# s.lstrip(rm)       删除s字符串中开头处，位于 rm删除序列的字符
# s.rstrip(rm)      删除s字符串中结尾处，位于 rm删除序列的字符
# 当rm为空时，默认删除空白符（包括'\n', '\r',  '\t',  ' ')
# =================
a='     tian'
a.strip()
a='\t\t\r\nabc'
a='tianyunchuan'
a.lstrip('t')
a.strip('t')
a='123abc'
a.strip('12')
a.strip('1ac')
a.strip('12ac')

####    6. extend
extend_list = []        # 必须先定义一个空列表
extend_list.extend([1])
extend_list.extend(["a", [3, 4] ])

l1 = [1,2,3]
l2 = [4,5,6,'a']
list_extend = []
list_extend.extend(l1)
list_extend.extend(l2)

####    7. find(str, pos_start, pos_end)
str_1 = "011213456"
str_1.find("63")
#1. if  str.find('23'):  此时默认为  str.find('23') != 0：
#2. find()函数找不到时返回为-1。
#函数原型：find(str, pos_start, pos_end)
#str:被查找“字串”
#pos_start:查找的首字母位置（从0开始计数。默认：0）
#pos_end: 查找的末尾位置（默认-1）
#返回值：如果查到：返回查找的第一个出现的位置。否则，返回-1。
str_1.find("2",1,-1)  
str_1.find("9",1,-1)  
str_1.find("0") 
str_1.find('4')

str1 = "this is string example....wow!!!";
str2 = "exam";

str1.find(str2)
str1.find(str2, 10)
str1.find(str2, 40)

info = 'abca'

info.find('a')
info.find('a',1)
info.find('3')


####    8. zip
x = [1,2,3]
y = [4,5,6]
z = [6,7,8]

xyz = zip(x,y,z)
xyz
list(xyz)
u = zip(*xyz)
#   zip、dict合用
word = ['me','you','he']
d = dict(zip(x,word))

####    9. [random]
import random
lst=list(range(1,11))  
random.choice(lst)      #随机取1个值
random.sample(lst,3)    #随机取n个值
random.shuffle(lst)     #打乱顺序
random.randint(1,1000)  #取随机数
random.random()         #取0-1间的浮点型随机数
random.getrandbits(10)    #n个比特位的随机数

random.seed()
print('随机数2：',random.random())
# 随机数一样
random.seed(1)
print('随机数3：',random.random())
random.seed(1)
print('随机数4：',random.random())
random.seed(2)
print('随机数5：',random.random())

# 抽取范围内随机数 序列
import random
samples_index = random.sample(range(0,60000),5) 

####    10. dir 当前文件变量清单
info=dir()

####    11. sys 设置最大递归数量
import sys
sys.setrecursionlimit(2000)


####    12. collections
## Counter
import collections
obj = collections.Counter('aabbccc')
print(obj)
d = dict(obj)
#输出：Counter({'c': 3, 'a': 2, 'b': 2})

s = ['a','b','a','c','d','d','a']
obj = collections.Counter(s)
d = dict(obj)

## ChainMap
baseline = {'music': 'bach', 'art': 'rembrandt'}
adjustments = {'art': 'van gogh', 'opera': 'carmen'}
list(collections.ChainMap(adjustments, baseline))

# Tally occurrences of words in a list
cnt = collections.Counter()      
for word in ['red', 'blue', 'red', 'green', 'blue', 'blue']:
    cnt[word] += 1
cnt


# Find the ten most common words in Hamlet
import re
#words = re.findall(r'\w+', open('hamlet.txt').read().lower())
words = re.findall(r'\w+', open('hamlet.txt').read().lower())
collections.Counter(words).most_common(10)


####    13. itertools
## 累加列表
x = itertools.accumulate(range(10))
print(list(x))
# [0, 1, 3, 6, 10, 15, 21, 28, 36, 45]

## 合并列表
x = itertools.chain(range(3),range(4), [3,2,1])
print(list(x))

## 笛卡尔积
x = itertools.product(['男', '女'], ['20代', '30代', '40代'])
print(list(x))

#### __str__
class Car():
    def __init__(self, name):
        self.name = name

    def __str__(self) -> str:
        text = f'这是一辆车,名字是{self.name}'
        return text

A = Car('BNW')
print(A)


print(dir(__builtins__))


""" sys """
#### 查看解释器路径 
import sys
print(sys.executable)

""" [MySQL] """
#pip install mysqlclient
#pip install mysql-python
#pip install PyMySQL


## __file__
#### 
print(os.__file__)


import MySQLdb
import numpy as np
import pandas as pd
from sqlalchemy import create_engine

## get data from MySQL
user_name = 'mlc_1001'
user_password = 'mill_metrix_9999'
db_name = '_tm_sales_bzy'
table_name = 'tm_sku_list_all'

connection = MySQLdb.connect(host='rm-uf675p1vvls0t85vko.mysql.rds.aliyuncs.com',user=user_name,passwd=user_password,port=3306,db=db_name,charset='utf8');
df_raw = pd.read_sql("SELECT * FROM {};".format(table_name),con=connection);

## 关闭
connection.close()

## DataFrame to MySQL
yconnect = create_engine('mysql+mysqldb://{}:{}@rm-uf675p1vvls0t85vko.mysql.rds.aliyuncs.com:3306/{}?charset=utf8'.format(user_name, user_password, db_name))
pd.io.sql.to_sql(df,'tm_sales_all', yconnect, schema=db_name, if_exists='append',index=False)
yconnect.dispose()



""" [if] [elif]"""
#  if 表达式 1: 
# 	语句块 1 
# elif表达式 2: 
# 	语句块 2 
# elif表达式 3: 
# 	语句块 3
# else: 
# 	语句块 n 




""" [捕获异常 不中断（异常基类，Exception）] """
## try except
try:
    sx003=(datetime.datetime.now()-datetime.datetime.strptime(sx002, "%Y-%m-%d")).days/365
except Exception:
    print('异常')
    x=100

try:
    sx003=(datetime.datetime.now()-datetime.datetime.strptime(sx002, "%Y-%m-%d")).days/365
except Exception:
    pass
    x=101


## try except continue
try:
    response = requests.get(url_httpbin, headers=headers, timeout=1, proxies=ip)
    if response.text != ip_home:
        
        r = requests.get(job_url, headers=headers, proxies=ip)
        html = r.content.decode('gbk')
        infos, d = get_infos(html)
        break

except OSError:
    pass
continue


## try finally
try:
    f = open('xx.jpg', 'rb')
    f.readlines()
finally:
    print('aaa')
    f.close()



""" [numpy] [np] """
import numpy as np
np.__version__

## numpy.array
np.where
np.percentile

nparr = np.array([i for i in range(10)])
nparr
type(nparr)     # > numpy.ndarray
# 调用
nparr[5]
# 赋值
nparr[5] = 100
nparr
nparr.dtype

nparr[5] = 333.14

nparr2 =np.array([1,2,4,5.666])
nparr2.dtype

nparr=[1,2,3.14]
nparr.dtype
np.zeros(10)
np.zeros(shape=(3,10),dtype=int)
np.ones(shape=(3,10 ),dtype=float)
arr=np.full(shape=(5,5),fill_value=999)

## 创建Numpy数组(和矩阵)
# 其他创建numpy.array的方法
nparr = np.zeros(10)
nparr.dtype

np.zeros(10,dtype=int)

nparr = np.zeros((3,5))
nparr = np.zeros(shape=(3,5),dtype=int)

np.ones((3,9))
np.full(shape=(2,5),fill_value=666.9)
n = np.random.permutation(100)

# arange
[i for i in range(0,20,2)]
nparr = np.arange(0,20,2)
nparr = np.arange(0,20,0.8)

np.arange(7)    

# linspace
np.linspace(0,20,4)     # 包括0，20
np.linspace(0,20,11)

# ramdom
np.random.randint(0,10,3)  #前闭后开
np.random.randint(0,100,size=3)    #前闭后开
np.random.randint(0,100,size=(3,9))    #前闭后开
np.random.seed(666)
np.random.randint(0,100,size=(3,9))  

np.random.random(10)
np.random.random((3,10))

#   array(倒序访问)
x = np.arange(10)
x1 = x[::-1]

#   reshape

X = np.arange(15).reshape(3,5)
X[:2,::2]

x=np.arange(0,9)
y=x.reshape(3,3)

# 展开
y.ravel()

#   二维反转
X[::-1,::-1]

# 扁平化数组、变为一维
x = np.array([[1, 2, 3], [4, 5, 6]])
np.ravel(x)
#-> array([1, 2, 3, 4, 5, 6])
#-> (6,)


#   permutation（乱序）
import numpy as np
n = np.random.permutation(100)
index = list(np.random.permutation(100))
max(n)
li = np.arange(0,100)
index = np.random.permutation(100)
li_02 = li[index]

## 正态分布随机数
np.random.normal()      # 均值为0，方差为1
np.random.normal(100,10,size=(3,4))  # 均值为100，方差为10

## Numpy数组(和矩阵)的基本操作
import numpy as np
# 基本操作
x = np.arange(10)
X = np.arange(15).reshape(3,5)
x.ndim
X.ndim
x.shape    # 返回元组
X.shape
x.size    # 元素总数量
X.size
# numpy.array的数据访问
x[0]
x[-1]
X[-1]       # 最后一行

X[2][2]    # 不建议
X[(2,2)]     # 建议！！！！
x[0:5]
x[:5]
x[5:]

x[::2]     # 步长
x[::-1]    # 倒序

X
X[:2,:3]
X[:2,::2]
X[::-1,::-1]    # 反转

X[0]
X[0,:]

X[:,0]    # 取列
X[:,0].ndim

subX = X[:2,:3]
subX[0,0]=100    # 影响原来矩阵
subX = X[:2,:3].copy()   # 不影响原数据

# Reshape
x.reshape((2,5))    # 不改变原数据
A = x.reshape((2,5))
B = x.reshape((1,10))

C = x.reshape((10,-1))    # -1 自动计算行、列数
C = x.reshape((-1,10))

x.reshape((2,-1))

## Numpy数组(和矩阵)的合并与分割
# 合并操作
x = np.array([1,2,3])
y = np.array([3,2,1])
z = np.array([666,666,666])
np.concatenate([x,y])    # 连接对象为 列表形式

A = np.arange(6).reshape((2,-1))
np.concatenate([A,A,A])
np.concatenate([A,A],axis=0)
np.concatenate([A,A],axis=1)
A2 = np.concatenate([A,z.reshape(1,-1)])

np.vstack([A,z])    # 智能判断唯独方向
B = np.full((2,2),100)
np.hstack([A,B])


# 合并
a = np.arange(9).reshape(3, 3)
b = (a + 1) * 10
# 横向合并 hstack
z = np.hstack((a, b))
# 纵向合并 vstack
z = np.vstack((a, b))

# 
z = np.dstack((a,b))

np.concatenate((a, b), axis = 0) # 总线
np.concatenate((a, b), axis = 1) # 横向

one_d_a = np.arange(5)
one_d_b = (one_d_a + 1) * 10
np.column_stack((one_d_a, one_d_b))
np.row_stack((one_d_a, one_d_b))
# 分割操作
x = np.arange(100)
x1,x2,x3 = np.split(x,[3,7])
np.split(x,[5,])
x1 ,x2 =np.split(x,[5,])

A = np.arange(20).reshape((5,4))
A1, A2 = np.split(A,[2])
A1, A2 = np.split(A,[2],axis=1)
upper,lower = np.vsplit(A,[2])
left, right = np.hsplit(A,[2])
data = np.arange(16).reshape((4,4,))
X, y = np.hsplit(data,[-1]) 

y[:,0]    # 转换为向量

## Universal Function
X = np.arange(1,16).reshape((3,5))
X + 1
X - 1
X * 2
X / 2
X // 2    # 整数除法
X ** 2
X % 2
1 / X
np.abs(X)
np.sin(X)
np.cos(X)
np.tan(X)
np.exp(X)      # e的X次方
np.power(3,X)
3 ** X
np.log(X)    # 自然对数
np.log2(X)    # 以2为底
np.log10(X)

## 矩阵运算 ！！！！
A = np.arange(4).reshape((2,2))
B = np.full((2,2),10)
A + B
A - B
A * B
A / B

A.dot(B)     # 矩阵运算（点乘）！！！！!！！！！！！！！！！！！！！！！！！！
A.T
A.T.dot(B)

# 向量和矩阵运算
v = np.array([1,2])
A = np.arange(4).reshape((2,2))
v + A
np.vstack([v]*A.shape[0])
np.vstack([v]*A.shape[0]) + A
np.tile(v,(2,1))
np.tile(v,(2,1)) + A
v * A
v.dot(A)
A.dot(v)

## 矩阵的逆
A = np.arange(4).reshape((2,2))s
invA = np.linalg.inv(A)
A.dot(invA)
invA.dot(A)
X = np.arange(8).reshape((2,4))
# invB = np.linalg.inv(B)    # 不是矩阵不能求，但是可以求伪逆
pinvX = np.linalg.pinv(X)    # 伪逆矩阵
X.dot(pinvX)
pinvX.dot(X)

# argsort
# https://blog.csdn.net/Python798/article/details/81138040
import numpy as np
x = np.array([3, 1, 2])
np.argsort(x)  
#> argsort函数返回的是数组值从小到大的索引值
#> array([1, 2, 0], dtype=int64)

## 通用函数
m = np.arange(10, 19).reshape(3, 3)
print (m)
print ("{0} min of the entire matrix".format(m.min()))
print ("{0} max of entire matrix".format(m.max()))
print ("{0} position of the min value".format(m.argmin()))
print ("{0} position of the max value".format(m.argmax()))
print ("{0} mins down each column".format(m.min(axis = 0)))
print ("{0} mins across each row".format(m.min(axis = 1)))
print ("{0} maxs down each column".format(m.max(axis = 0)))
print ("{0} maxs across each row".format(m.max(axis = 1)))

a = np.arange(1,10)
a.mean(), a.std(), a.var()
a.sum(), a.prod()
a.cumsum(), a.cumprod()

## nan定义为0
import numpy as np
np.set_printoptions(precision=8)
x = np.array([np.inf, -np.inf, np.nan, -128, 128])
np.nan_to_num(x)
array([  1.79769313e+308,  -1.79769313e+308,   0.00000000e+000,
        -1.28000000e+002,   1.28000000e+002])



















""" [DataFrame] """
import pandas as pd, numpy as np

## 导入excel设置第一列为index
df = pd.read_excel(r'C:\Users\tianyunchuan\iCloudDrive\Desktop\_pyCodeSpyder_\_data\data_survey_proc\ca_mca_nabe.xlsx', index_col=0)


df3=pd.DataFrame(np.array([[10,20],[30,40]]),index=['a','b'],columns=['c1','c2'])
df3.index
df3.columns
# 列数值
df3.columns[1]
df3.info()

eu12=pd.read_csv(r'C:\_TianProc\_data_\data_class\Python_basic_class_163_zjd_data\Eueo2012.csv',index_col="Team")
eu12_goals = eu12.copy()

eu12.shape
eu12.index
eu12.columns

#### 选择列
eu12.iloc[:,:2]
eu12[[1,2]] #失败，因为列名已有定义

## 获取所爱第几列
eu12.columns.get_loc('Red Cards')


####    选择行
eu12[:5]
eu12['Croatia':'Denmark']
eu12.loc['Denmark']
eu12.iloc[[2,4]]
eu12.index.get_loc('France')
eu12.at['France','Goals']
eu12[2:4][['Goals','Shots on target']]

#### 行号取行
df.loc[1,3,5]
df.loc[list(range(0,18)) + list(range(34,37))]

df.iloc[[1,3],[2,4,6]]

# DataFrame插入行
eu12_goals.columns
eu12_goals.insert(3,'tian',eu12_goals.Touches*100)
df3.insert(0,'tian','ok')


####    布尔选择
eu12_goals=eu12.copy()
eu12.Goals>4
eu12_goals[(eu12_goals.Goals>4)&(eu12_goals.s_on_targets>15)]


## 添加行
df1=eu12_goals.iloc[:5]
df2=eu12_goals.iloc[[10,11,12]]
df3=df1.append(df2)

df4=pd.DataFrame(2,index=df1.index,columns=['changes'])
df5 = df1.append(df4)

df6 = df1.append(df4,ignore_index=True)  #忽略index,直接添加

## 删除行、删除列
data = {'name': ['Jason', 'Molly', 'Tina', 'Jake', 'Amy'], 
        'year': [2012, 2012, 2013, 2014, 2014], 
        'reports': [4, 24, 31, 2, 3]}
df = pd.DataFrame(data, index = ['Cochice', 'Pima', 'Santa Cruz', 'Maricopa', 'Yuma'])
df

## 删除列 第几列 第n列
df = df.drop(df.columns[[0,3,6]], axis = 1)

# 删除行
df.drop(['Cochice', 'Pima'],inplace=True)
# 删除列
df.drop('reports', axis=1, inplace=True)

## concat
df = pd.concat([df1,df2])
concat(objs, axis=0, join='outer', join_axes=None, ignore_index=False, keys=None, levels=None, names=None, verify_integrity=False, sort=None, copy=True)

## set_index
sp500=pd.read_csv(r'C:\Users\tianyunchuan\iCloudDrive\Desktop\_jupyter_\Python_basic_class_163_zjd\data\sp500.csv')

sp500 = sp500.set_index("Symbol")
sp500.loc['ACE']

## reset_index
sp500=sp500.reset_index()
df_ds.reset_index(drop=True, inplace=True)

## multi_index
np.arrays = [['one','one','one','two','two','two'],[1,2,3,1,2,3]]

df = pd.DataFrame(np.random.randn(6,2),index=pd.MultiIndex.from_tuples(list(zip(*np.arrays))),columns=['A','B'])
df.loc["one"]
df.loc["one"].loc[1]


df.set_index(['school','class'])

## 两个列表创建DataFarame
import numpy as np
s1 = [1,2,3,4]
s2 = ['a', 'b','c','d']
df = pd.DataFrame({'name':s1, 'number':s2})


## 读excel工作表名生成列表
import pandas as pd

# 读取Excel文件
excel_file = pd.ExcelFile('your_excel_file.xlsx')

# 获取所有工作表的名称
sheet_names = excel_file.sheet_names

# 创建一个新的DataFrame来存储工作表名
df = pd.DataFrame({'编号': range(1, len(sheet_names) + 1), '工作表名': sheet_names})

# 将数据写入新的Excel文件
df.to_excel('output_excel_file.xlsx', index=False)



## set_index + mean, sum
import pandas as pd
df = pd.read_excel(r'C:\Users\tianyunchuan\iCloudDrive\Desktop\_pyCodeSpyder_\_data\data_training\4.2 OE_Message Playback.xlsx')

df1 = df.set_index("city age")
df1.sum(level='city age')  
result1 = df1.sum(level='city age')
result1 = df1.mean(level='city age')

df2 = df.set_index(['city age', 'gender'])
result2 = df2.sum(level='city age')  

result2 = df2.sum(level=['city age','gender'])

## sort_index
r2 = result1.sort_index(ascending=True)#此时为index降序    

z = df['city age'].unique() 
sp500.sort_index(ascending=False)

## sort_value
eu12.sort_values(['Red Cards', 'Yellow Cards'], ascending = False)[['Red Cards', 'Yellow Cards']]
eu12.sort_values(['Red Cards', 'Yellow Cards'], ascending = True)[['Red Cards', 'Yellow Cards']]


import pandas as pd
df = pd.read_excel(r'C:\Users\tianyunchuan\iCloudDrive\Desktop\_pyCodeSpyder_\_data\data_training\country.xlsx')

## 计算字段Nan数量
train_titanic.isna().sum()



## 返回Numpy值数组
import pandas
data_pd = pd.read_csv('chengdu.csv', header=0, index_col=0)
data_values = data_pd.values

# 返回最大最小值的index
df.describe()
df[['Population','GNP']].idxmin()
df[['Population','GNP']].idxmax()

s = df['Continent'].count()
df['Continent'].describe()

####   0. 获取数据（From Mongo）
df = pd.DataFrame(infos)

import _utils_.DB_MySql.Get_Data_From_MySql as getSQL
getSql = getSQL.Get_Data_From_MySql('world', 'country')
df = getSql.go()

####    1. 调用1列为Series (DataFrame→Series) 
df1=df['Name']

####    2. 调用1列为list (DataFrame→list) 
df1=list(df['Name'])

####    3. 调用1列为DataFrame (DataFrame→DataFrame) 
df1=df[['Name']]
df1=df.loc[100:120,['Name']]

####    4. 调用多列 (DataFrame→DataFrame)
df1=df[['Code','Name','GNP']]

####    5. 时间转换（str>datetime）
df['purchase_time'] = df['订单付款时间 '].apply(lambda x: datetime.datetime.strptime(x,'%Y/%m/%d %H:%M')) 

import time
 
# 将时间字符串转为时间戳int
dt = "2016-05-05 20:28:54"
# 转换成时间数组
timeArray = time.strptime(dt, "%Y-%m-%d %H:%M:%S")
# 转换成时间戳
timestamp = time.mktime(timeArray)
 
print(timeArray)
print("时间戳-timestamp: %s" % timestamp)

## 根据时间筛选数据
df_18 = df[(df['订单付款时间'] > '2018-01-01 00:0:00') & (df['订单付款时间'] < '2019-01-01 00:00:00')]

####    插入指定任意位置
map_all.insert(0,'question_no_ALL','')    # 0>位置、'question_no'>列名、'内容'

####    插入列
col_name = data_all.columns.tolist()
col_name.insert(col_name.index('Q_00001_N'),'survey_series_num')# 在某列前面列前面插入
data_all = data_all.reindex(columns=col_name)

#####   loc方式
df1=df.loc[90:100,['Name']]
df1=df.loc[90:100,['Code','Name','GNP']]
df1=df.loc[:,['Code','Name','GNP']]


##    条件调用DataFrame （单条件完全匹配）
df1=df[df.GNP>500]
df1=df[df.Continent =='Asia']

##       isin
alist=['Asia','Europe']
df1=df[df['Continent'].isin(alist)]
df1=df[df['Continent'].isin(['Asia','Europe'])]

##    条件调用DataFrame （多条件完全匹配）
df1=df[(df['Continent'] =='Asia') & (df['GNP'] >10000)]     # and
df1=df[(df['Continent'] =='Asia') | (df['GNP'] >10000)]     # or
df1=df[~(df['Continent'] =='Asia') | (df['GNP'] >10000)]     # ~取反

##      取反~~~~~~~~
df1=df[~df['Continent'].isin(['Asia','Europe'])]

##    修改列名
#http://www.cnblogs.com/hhh5460/p/5816774.html
# https://blog.csdn.net/liuweiyuxiang/article/details/78445440
# ④暴力（好处：也可只修改特定的列）
df.rename(columns={'$a': 'a', '$b': 'b', '$c': 'c', '$d': 'd', '$e': 'e'}, inplace=True) 

##    条件调用DataFrame （单条件like）
df1=df[df['Continent'].str.contains(r'.*?A.*')]
df1=df[df['Continent'].str.contains(r'.*?(As|Af).*')]
df1=df[df['Continent'].str.contains(r'.*?Asi.*')]
df2=df[df['Continent'].str.contains(r'A')]


##  diff函数
import pandas as pd
df = pd.DataFrame( {'a':[1,2,3,4,5],
                    'b':[6,7,8,9,10],
                    'c':[11,12,13,14,15]})

# axis=0或index表示上下移动， periods表示移动的次数，为正时向下移，为负时向上移动。
df.diff(periods=1, axis='index')    # 下一行减去前一行
df.diff( periods=-1, axis=0)        # 前行 减去 下一行

df.diff( periods=1, axis='columns')
df.diff( periods=2, axis='columns')

df.diff( periods=-1, axis=1)

# https://blog.csdn.net/qq_32618817/article/details/80653841



##    判断 [nan]
import math
x=float('nan')



x
math.isnan(x)
# >>三元True

# 读取excel文件时 空值不要导入为[nan]
map_all = pd.read_excel('{}\{}.xlsx'.format(PATH, MAP_FILE_NAME), sheet_name='variable', ignore_index=True, keep_default_na=False)

if str(s_cell)=='nan':
    print(str(s_cell))



##    空字符替换成Nan
import numpy as np
df1=df.loc[:,:].replace(r'','Nan') 
df2=df.loc[:,:].replace(r'',np.NaN) 

## 0替换成Nan
df.replace(0,'np.nan', inplace=True)

##    删除包含Nan的行
newDF = df.dropna()
newDF = df.dropna(subset=['Capital'])
newDF = df.dropna(subset=['Capital','IndepYear'])

#   DataFrame 删除某列含空白的行
df = df[~((df.Name) =='')]
df.Name

#   判断空值，col是列名称 isnull()
df.Name.isnull()
df.A.isnull()
sum(df.A.isnull())

##  计算个列 nan的个数
import pandas as pd
import numpy as np
from numpy import nan as NaN

df = pd.read_csv(r'C:\_data\ml_163_master\data\iris-data.csv')
df.head()
df.isnull().values.any()
df.isnull().sum()
newDF = df.dropna()
df = pd.read_csv(r'C:\_data\ml_163_master/data/iris-data.csv', na_values=['NA']) #pandas缺省对NA已经处理了

#### [fillna()函数详解]
## https://blog.csdn.net/weixin_39549734/article/details/81221276
#inplace参数的取值：True、False
#method参数的取值 ： {‘pad’, ‘ffill’,‘backfill’, ‘bfill’, None}, default None
#pad/ffill：用前一个非缺失值去填充该缺失值
#backfill/bfill：用下一个非缺失值填充该缺失值
#None：指定一个值去替换缺失值（缺省默认这种方式）
#limit参数：限制填充个数
#axis参数：修改填充方向
#df = pd.DataFrame([[np.nan,2,np.nan,0],[3,4,np.nan,1],[np.nan,np.nan,np.nan,5],[np.nan,3,np.nan,4]],columns=list('ABCD'))

df=pd.DataFrame([[1,2,3],[NaN,NaN,2],[NaN,NaN,NaN],[8,8,NaN]],columns=list('ABC'))
# 将NAN值转换为0
df.fillna(0)
# 向前或向后传播
df.fillna(method='ffill')
df.fillna(method='bfill')
# 用字典填充
df.fillna({'A':10, 'B':20, 'C':30})
# 指定inplace参数
df.fillna(0,inplace=True))
# 指定limit参数
df.fillna(method='bfill', limit=2)
# 指定axis参数
df.fillna(method="ffill", limit=1, axis=1)

data['灰度分'] = data['灰度分'].fillna(data['灰度分'].mean()))
data['灰度分'] = data['灰度分'].fillna(data['灰度分'].mode()【0))

train_titanic['Age'] = train_titanic['Age'].fillna(train_titanic['Age'].median())
train_titanic['Embarked'] = train_titanic['Embarked'].fillna(train_titanic['Embarked'].mode()[0])
# https://blog.csdn.net/vivian_ll/article/details/91900323
from fancyimpute import KNN

data = pd.read_csv(path,encoding='gbk')
data = pd.DataFrame(KNN(k=6).fit_transform(data)) 
data.columns = ['sex','age','label']  # fancyimpute填补缺失值时会自动删除列名


# ！！！这里使用的是fancyimpute库，安装的时候需要visual C++环境。

# KNN预测的步骤是选择出其他不存在缺失值的列，同时去除掉需要预测缺失值的列存在缺失值的行，然后计算距离。
# 如果缺失值是离散的，使用K近邻分类器，投票选出K个邻居中最多的类别进行填补；如果为连续变量，则用K近邻回归器，拿K个邻居中该变量的平均值填补。




####   nan处理 
#   https://blog.csdn.net/brucewong0516/article/details/80406564
#   nan填充为*
df = df.fillna(value='*')

## isin（df操作） 多条件筛选
df1 = df[~df['shop'].isin(['专营店','专卖店','阿里健康大药房'])]
set(df1['shop'])
df1 = df[df['shop'].isin(['专营店','专卖店'])]

## 包含多个文本（or）
import re
s='这个包装看上去时髦'
words=['时髦','流行','时尚']
pat=re.compile('|'.join(words))
1 if pat.search(s) else 0 
0 if pat.search(s)==None else 1


## categories
df['价格贵1']=df['价格贵'].astype('category')
df['价格贵1'].cat.categories=['male','female']  #即将0，1先转化为category类型再进行编码。


## astype("category").cat.codes
data['AIRLINE']
data['AIRLINE'].astype("category").cat.codes
c = data['AIRLINE'].astype("category")
d = dict(enumerate(c.cat.categories))d
# {0: 'AA',
#  1: 'AS',
#  2: 'B6',
#  3: 'DL',
#  4: 'EV',
#  5: 'F9',}

## if_else + 迭代[] + float
df['月成交1']=[(float(s.replace('万', ''))*10000 if '万' in s else s) for s in df.月成交]

## 排序 sort_values
df1 = df.sort_values(by='s_qty', axis=0, ascending=False, inplace=False, kind='quicksort', na_position='last')

## 去重
df.drop_duplicates(subset=['Continent'], keep='first', inplace=True)
df1=df.drop_duplicates(subset=['Continent'], keep='last', inplace=False)

data.drop_duplicates(subset=['A','B'],keep='first',inplace=True)
#subset： 列名，可选，默认为None#
#keep： {‘first’, ‘last’, False}, 默认值 ‘first’
#    first： 保留第一次出现的重复行，删除后面的重复行。
#    last： 删除重复项，除了最后一次出现。
#    False： 删除所有重复项。

## 条件生成新列
df["利润率评价"] = ''
df.loc[df["利润"]/df["销售额"]>0.15, "利润率评价"] = "高利润"
df.loc[df["利润"]/df["销售额"]<=0.15, "利润率评价"] = "低利润"


## explode
df = pd.DataFrame({'A': [[0, 1, 2], 'foo', [], [3, 4]],
                   'B': 1,
                   'C': [['a', 'b', 'c'], np.nan, [], ['d', 'e']]})
df.explode('A')


## DataFrame 根据元素查找index
df = pd.DataFrame({'BoolCol': [1, 2, 3, 3, 4],'attr': [22, 33, 22, 44, 66]},  
       index=[10,20,30,40,50])  
print(df)  
a = df[(df.BoolCol==3)&(df.attr==22)].index.tolist()  
print(a)  

## sort_values
DataFrame.sort_values(by='売上げ総額', axis=0, ascending=True, inplace=False, kind='quicksort', na_position='last')

## DataFrame完整参数
from pandas import DataFrame
import pandas as pd
import numpy as np
df=DataFrame(np.random.randn(6,4),columns=list('abcd'),index=['bj','sh','gz','sy','wh','nj'])
df=DataFrame(np.arange(12).reshape(3,4),columns=list('ABCD'),index=['a0','a1','a2'])



## replace
df['address_01'].replace(_k, _v,inplace=True)

## 行列重命名
df.rename(columns={'A':'A_rename'})  
df.rename(index={'a1':'other'})

df.rename(columns={'B':'CB'})

## 计算同一个column相邻两个数字之间的变化率
c1=DataFrame(np.arange(12).reshape(3,4),columns=list('ABCD'),index=['a0','a1','a2'])
c1['D'].pct_change() 

## 替换DataFrame中的特定数值,例如:将数据中的所有的1和2都替换成3
df.replace([1,2,11], 'aa', inplace=True)
c1.replace([1,2,11], 'aa', inplace=True)
c1.replace(['aa'], 'bb', inplace=True)

## PANDAS 数据合并与重塑（concat篇）
# https://blog.csdn.net/stevenkwong/article/details/52528616

## 获取列名
result.columns.values
list(result.columns.values)

## Dataframe数据上下合并
weibo_personal=pd.concat([weibo_personal,df_0],axis=0,ignore_index=True, sort=False)

##  建立空dataframe
df_tmp = pd.DataFrame()

df = pd.DataFrame(columns=['title', 'href', 'price', 'shop', 's_qty', 'c_qty', 'cate','timer', 'skuId'])  

dic1={'name':['小明','小红','狗蛋','铁柱'],'age':[17,20,5,40],'gender':['男','女','女','男']}
df3=pd.DataFrame(dic1)

##    删除行
df1=df.drop(4)
df.drop(['Cochice', 'Pima'])

##    删除单列

del df['城市']
df['城市']='上海'

df1=df.drop('城市',1)
df1=df.drop('城市',axis=1)
df.drop('当前时间',axis=1, inplace=True)


##    删除多列
df1=df.drop(['城市','当前时间'],axis=1)

##    追加列
df['城市']='上海'

##      dictToDataFrame (dict转DataFrame)
df = pd.DataFrame.from_dict(anchors,columns=['title','href','price','shop','s_qty','c_qty','cate'])

##      dataFrame 根据内容文本找出index索引
import pandas as pd
df = pd.DataFrame({'BoolCol': [1, 2, 3, 3, 4],'attr': [22, 33, 22, 44, 66]},index=[10,20,30,40,50])
print(df)
a = df[(df.BoolCol==3)&(df.attr==22)].index.tolist()
print(a)

## [merge]
# https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.merge.html

import  pandas as pd
import numpy  as np
df1 = pd.DataFrame({'lkey': ['foo', 'bar', 'baz', 'foo'],'value': [1, 2, 3, 5]})
df1
df2 = pd.DataFrame({'rkey': ['foo', 'bar', 'baz', 'foo'],'value': [5, 6, 7, 8]})
df2
df1 = df1.drop([2])

df1.merge(df2)
df1.merge(df2,left_on="lkey",right_on="rkey")

df = df1.merge(df2, left_on='lkey', right_on='rkey',suffixes=('_left', '_right'))


https://blog.csdn.net/bin083/article/details/94978218
#只在table1添加分数列，不添加排名列
result = pd.merge(df1,df2.loc[:,['学号','分数']],how='left',on = '学号')

## rolling
import pandas as pd
import numpy as np
df = pd.DataFrame({'1': ['A1', 'A2', 'A1', 'A2', 'A2', 'A1', 'A2'],
     '2': ['B1', 'B1', 'B1', 'B1', 'B1', 'B1', 'B1'],
     'num': [1,2,1,3,4,2,1]}, 
     index = [pd.Timestamp('20130101 09:00:00'),
       pd.Timestamp('20130101 09:00:01'),
       pd.Timestamp('20130101 09:00:02'),
       pd.Timestamp('20130101 09:00:03'),
       pd.Timestamp('20130101 09:00:04')
       pd.Timestamp('20130101 09:00:05'),
       pd.Timestamp('20130101 09:00:06')])

df.groupby(['1', '2'])['num'].rolling('3s').sum()


df = pd.DataFrame({'B': [0, 1, 2, np.nan, 4]})
df.rolling(2, win_type='triang').sum()
df.rolling(2, win_type='gaussian').sum(std=3)
df.rolling(2).sum()
df.rolling(2, min_periods=1).sum()

indexer = pd.api.indexers.FixedForwardWindowIndexer(window_size=2)
df.rolling(window=indexer, min_periods=1).sum()


## pd.shift
import pandas as pd
import datetime
df = pd.DataFrame(np.arange(16).reshape(4,4),columns=['AA','BB','CC','DD'],index =pd.date_range('6/1/2012','6/4/2012'))

## 删除Unnamed列
df = df.loc[ : , ~df.columns.str.contains("^Unnamed")]


df
Out[38]: 
            AA  BB  CC  DD
2012-06-01   0   1   2   3
2012-06-02   4   5   6   7
2012-06-03   8   9  10  11
2012-06-04  12  13  14  15

df.shift(freq=datetime.timedelta(1))
Out[39]: 
            AA  BB  CC  DD
2012-06-02   0   1   2   3
2012-06-03   4   5   6   7
2012-06-04   8   9  10  11
2012-06-05  12  13  14  15

df.shift(freq=datetime.timedelta(-2))
Out[40]: 
            AA  BB  CC  DD
2012-05-30   0   1   2   3
2012-05-31   4   5   6   7
2012-06-01   8   9  10  11
2012-06-02  12  13  14  15



df = pd.DataFrame(np.arange(16).reshape(4,4),columns=['AA','BB','CC','DD'],index =['a','b','c','d'])

df
Out[14]: 
   AA  BB  CC  DD
a   0   1   2   3
b   4   5   6   7
c   8   9  10  11
d  12  13  14  15
#当period为正时，默认是axis = 0轴的设定，向下移动
df.shift(2)
Out[15]: 
    AA   BB   CC   DD
a  NaN  NaN  NaN  NaN
b  NaN  NaN  NaN  NaN
c  0.0  1.0  2.0  3.0
d  4.0  5.0  6.0  7.0
#当axis=1，沿水平方向进行移动，正数向右移，负数向左移
df.shift(2,axis = 1)
Out[16]: 
   AA  BB    CC    DD
a NaN NaN   0.0   1.0
b NaN NaN   4.0   5.0
c NaN NaN   8.0   9.0
d NaN NaN  12.0  13.0
#当period为负时，默认是axis = 0轴的设定，向上移动
df.shift(-1)
Out[17]: 
     AA    BB    CC    DD
a   4.0   5.0   6.0   7.0
b   8.0   9.0  10.0  11.0
c  12.0  13.0  14.0  15.0
d   NaN   NaN   NaN   NaN

## df.pct_change
# https://pandas.pydata.org/pandas-docs/version/0.23.4/generated/pandas.DataFrame.pct_change.html
# diff
# df[‘column’].diff(-1)计算column列这一行减去上一行数据之差

# pct_change
# df[‘column’].pct_change(-1)计算column列这一行减去上一行数据之差再除以上一行，即每日涨跌幅
s = pd.Series([90, 91, 85])
s
0    90
1    91
2    85
dtype: int64

s.pct_change()
0         NaN
1    0.011111
2   -0.065934
dtype: float64


s.pct_change(periods=2)
0         NaN
1         NaN
2   -0.055556
dtype: float64


df = pd.DataFrame({
...     'FR': [4.0405, 4.0963, 4.3149],
...     'GR': [1.7246, 1.7482, 1.8519],
...     'IT': [804.74, 810.01, 860.13]},
...     index=['1980-01-01', '1980-02-01', '1980-03-01'])
df
                FR      GR      IT
1980-01-01  4.0405  1.7246  804.74
1980-02-01  4.0963  1.7482  810.01
1980-03-01  4.3149  1.8519  860.13

df.pct_change()


df = pd.DataFrame({
...     '2016': [1769950, 30586265],
...     '2015': [1500923, 40912316],
...     '2014': [1371819, 41403351]},
...     index=['GOOG', 'APPL'])
df
          2016      2015      2014
GOOG   1769950   1500923   1371819
APPL  30586265  40912316  41403351

df.pct_change(axis='columns')
      2016      2015      2014
GOOG   NaN -0.151997 -0.086016
APPL   NaN  0.337604  0.012002

## df_date_range
import pandas as pd
dates = pd.date_range(start="2018-10-01", end="2018-10-12")
dates1 = pd.date_range(start="2018-10-01", periods=6)
print(dates)
print(dates1)
DatetimeIndex(['2018-10-01', '2018-10-02', '2018-10-03', '2018-10-04',
               '2018-10-05', '2018-10-06', '2018-10-07', '2018-10-08',
               '2018-10-09', '2018-10-10', '2018-10-11', '2018-10-12'],
              dtype='datetime64[ns]', freq='D')
DatetimeIndex(['2018-10-01', '2018-10-02', '2018-10-03', '2018-10-04',
               '2018-10-05', '2018-10-06'],
              dtype='datetime64[ns]', freq='D')


import pandas as pd
dates = pd.date_range(start="2018-10-01", periods=8, freq="B")
dates1 = pd.date_range(start="2018-10-01", periods=8, freq="H")
dates2 = pd.date_range(start="2018-10-01", periods=8, freq="S")



import pandas as pd
dates = pd.date_range(start="2018-10-01 03:00:00", periods=8, freq="H", normalize=True)
dates1 = pd.date_range(start="2018-10-01 03:00:00", periods=8, freq="H", normalize=False)


## resample
# 首先创建一个Series，采样频率为一分钟。
index = pd.date_range('1/1/2000', periods=9, freq='T')
series = pd.Series(range(9), index=index)
series
2000-01-01 00:00:00    0
2000-01-01 00:01:00    1
2000-01-01 00:02:00    2
2000-01-01 00:03:00    3
2000-01-01 00:04:00    4
2000-01-01 00:05:00    5
2000-01-01 00:06:00    6
2000-01-01 00:07:00    7
2000-01-01 00:08:00    8
Freq: T, dtype: int64


# 降低采样频率为三分钟
series.resample('3T').sum()
2000-01-01 00:00:00     3
2000-01-01 00:03:00    12
2000-01-01 00:06:00    21
Freq: 3T, dtype: int64

# 降低采样频率为三分钟，但是每个标签使用right来代替left。请注意，bucket中值的用作标签。
series.resample('3T', label='right').sum()
2000-01-01 00:03:00     3
2000-01-01 00:06:00    12
2000-01-01 00:09:00    21
Freq: 3T, dtype: int64

# 降低采样频率为三分钟，但是关闭right区间。
series.resample('3T', label='right', closed='right').sum()
2000-01-01 00:00:00     0
2000-01-01 00:03:00     6
2000-01-01 00:06:00    15
2000-01-01 00:09:00    15
Freq: 3T, dtype: int64


# 增加采样频率到30秒
series.resample('30S').asfreq()[0:5] #select first 5 rows
2000-01-01 00:00:00     0
2000-01-01 00:00:30   NaN
2000-01-01 00:01:00     1
2000-01-01 00:01:30   NaN
2000-01-01 00:02:00     2
Freq: 30S, dtype: float64

# 增加采样频率到30S,使用pad方法填充nan值。
series.resample('30S').pad()[0:5]
2000-01-01 00:00:00    0
2000-01-01 00:00:30    0
2000-01-01 00:01:00    1
2000-01-01 00:01:30    1
2000-01-01 00:02:00    2
Freq: 30S, dtype: int64

# 增加采样频率到30S,使用bfill方法填充nan值。
series.resample('30S').bfill()[0:5]
2000-01-01 00:00:00    0
2000-01-01 00:00:30    1
2000-01-01 00:01:00    1
2000-01-01 00:01:30    2
2000-01-01 00:02:00    2
Freq: 30S, dtype: int64


# 通过apply运行一个自定义函数
def custom_resampler(array_like):
...     return np.sum(array_like)+5
series.resample('3T').apply(custom_resampler)
2000-01-01 00:00:00     8
2000-01-01 00:03:00    17
2000-01-01 00:06:00    26
Freq: 3T, dtype: int64

## df series 去重
df.iloc[:,i].unique()
df['cloName'].unique()
df['cloName'].nunique()  #查看数量

































""" [Series] """
import pandas as pd
# Base
price = pd.Series([15280,45888,15692,55689,28410,27566])
price = pd.Series([15280,45888,15692,55689,28410,27566],name="price")

temp = {'Mon': 33, 'Tue': 19, 'Wed': 15, 'Thu': 89, 'Fri': 11, 'Sat': -5, 'Sun': 9}
pd.Series(temp)

price.mean()
price.sum()
price.head(2)
price.tail(3)

dir(price)

# dtype
price = pd.Series([15280,45888,15692,55689,28410,27566.3])
price.dtype

temp=pd.Series([{},[],(3,4)])
temp.dtype
# -> Out[48]: dtype('O')

x=pd.Series(['2016-01-01','2017-01-01'])
x.dtype
# -> Out[62]: dtype('O')

x=pd.Series(['2016-01-01','2017-01-01'])
x = pd.to_datetime(x)
x.dtype

# dtype category
x=pd.Series(['hw','apple','vivo','mi','hw','oppo','samsung','vivo'],dtype='category')
x

dir(temp)
temp.__dir__()


## index
# index 是可以重复的！！！！！
import pandas as pd
price = pd.Series([15280,45888,15692,55689,28410,27566.3],index=['wh','sh','hz','bj','gz','nj'])

price['sh']
price.index
price.index.name='city'

# 删除
del price['sh']

# 时间分割
dates=pd.date_range('2016-01-01','2016-6-01',freq='M')
tempature=pd.Series([13,15,20,27,29],index=dates)

temp=pd.Series([13,15,20,27,29],index=[0,2,2,3,4])


## index/slicing
temp=pd.Series([13,15,20,27,29])
price.append(pd.Series([9500],index=['cd']))  #return a new series
price.set_value('bj',62000) # update inplace

price.min(), price.max(), price.median(), 
price.describe()
price.quantile(0.25)
price.describe(percentiles=[0.25,0.5,0.75])

## Series相加 自动对齐
s=pd.Series([10,20,30,40])
s2=pd.Series([10,20,30],index=[2,3,4])
s+s2

## Series.apply(---)
s.apply(float)

for num in s:
    print(num)

40 in s.values # 获取值
'bj' in price  # 获取index

## series 操作 和 Dict很接近
# looping over dictionary keys and values
for k,v in price.items():
    print(k,v)

temp=pd.Series(['hw','apple','vivo','mi','hw','oppo','samsung','vivo'],dtype='category')
temp.value_counts()



## Series→list        
c1=df['人均'].tolist()
c1=list(df['人均'])


## cat 字符串拼接
from pandas import Series
x = [1,2,3]   #  报错，数字类型
x = ['1', '2', '3']
word = ['me','you','he']

Series(x).str.cat(word, sep='-')

S = Series(['a', 'b', 'c']).str.cat(['A', 'B', 'C'], sep=',')
S = Series(['a', 'b', 'c']).str.cat(['A', 'B', 'C'], sep='-')

S = Series(['a', 'b', 'c']).str.cat(sep=',')

Series(['a', 'b']).str.cat([['x', 'y'], ['1', '2']], sep=',')

# split() 切分字符串
#http://www.runoob.com/python/att-string-split.html
import pandas as pd
import numpy as np
s = pd.Series(['a_b_c_e_f', 'c_d_e_f_g', np.nan, 'f_g_h_i_j_k'])
s.str.split('_')
s.str.split('_', -1)
s.str.split('_', 0)
s.str.split('_', 1)
s.str.split('_', 2)
s.str.split('_', 3)
s.str.split()
'a_b_c_e_f'.split('_')

# repeat
s = pd.Series(['a_b_c_e_f', 'c_d_e_f_g', np.nan, 'f_g_h_i_j_k'])
s.str.repeat(3)

df.Name.str.repeat(3)
df.Name.repeat(3)

# join() 对每个字符都用给点的字符串拼接起来，不常用
import numpy as np
s = pd.Series(['a_b_c_e_f', 'c_d_e_f_g', np.nan, 'f_g_h_i_j_k'])
s.str.join("!")
df.cate.str.join('!!')

str1 = "-";
seq = ("a", "b", "c"); # 字符串序列
print (str1.join( seq ));
','.join(['a','aaa'])		

# contains
s = pd.Series(['a_b_c_e_f', 'c_d_e_f_g', np.nan, 'f_g_h_i_j_k'])
type(s)
s.str.contains('d')
# sum(df['shop'].str.contains('旗舰'))

# replace() 替换
s = pd.Series(['a_b_c_e_f', 'c_d_e_f_g', np.nan, 'f_g_h_i_j_k'])
    s.str.replace("_", ".")
    'wolrd'.replace('d','_')
a.replace('\n', '').replace('\r', '')  #替换换行符！！
# 补齐
s = pd.Series(['a_b_c_e_f', 'c_d_e_f_g', np.nan, 'f_g_h_i_j_k'])
s.str.pad(10, fillchar=" ")
s.str.pad(10, fillchar="0")
s.str.pad(10, side="right", fillchar="?")
s.str.center(10, fillchar="?")
s.str.ljust(10, fillchar="?")
s.str.rjust(10, fillchar="?")
s.str.zfill(10)

## 填充和对齐包含
#^<>分别表示居中、左对齐、右对齐，后面带宽度
'{:>10}'.format('jihite')
'{:<10}'.format('jihite')
'{:^10}'.format('jihite')

'{:>30}'.format('jihite')
'{:<30}'.format('jihite')
'{:^30}'.format('jihite')


#  wrap() 在指定的位置加回车符号
s = pd.Series(['a_b_c_e_f', 'c_d_e_f_g', np.nan, 'f_g_h_i_j_k'])
s.str.wrap(3)

# slice() 按给定的开始结束位置切割字符串
s = pd.Series(['a_b_c_e_f', 'c_d_e_f_g', np.nan, 'f_g_h_i_j_k'])
s.str.slice(1,3)
df['shop'].str.slice(1,5)

# slice_replace() 使用给定的字符串，替换指定的位置的字符
s = pd.Series(['a_b_c_e_f', 'c_d_e_f_g', np.nan, 'f_g_h_i_j_k'])
s.str.slice_replace(1, 3, "?")
df['shop'].str.slice_replace(0,2,'T')

# 计算给定单词出现的次数
s = pd.Series(['a_b_ac_e_f', 'c_d_e_f_g', np.nan, 'af_g_h_i_j_k'])
s.str.count("a")
df['brand'].str.count('combi')

# startswith() endswith() 判断是否以给定的字符串开头
s1=s.str.startswith("a")
df1=df['brand'].str.startswith("a")
s.str.endswith("e")
s.str.endswith("a")
s.str.startswith("a")

####                27. findall() 查找所有符合正则表达式的字符，以数组形式返回
s = pd.Series(['a_b_c_e_f', 'c_d_e_f_g', np.nan, 'f_g_h_i_j_k'])

s1=s.str.findall("[d-z]");
df['shop'].str.findall("[旗]")

#括号匹配、括号查找
last_page=re.findall(re.compile(r'共(\d+)页'),first_htmls)

####                28. match() 检测是否全部匹配给点的字符串或者表达式
s1=s.str.match("[d-z]");
df['shop'].str.match("[旗]")
####                29. extract() 抽取匹配的字符串出来，注意要加上括号，把你需要抽取的东西标注上
s1=s.str.extract("([e-z])");
df['shop'].str.extract("([旗])")

# partition() 把字符串数组切割称为DataFrame，注意切割只是切割称为三部分，分隔符前，分隔符，分隔符后
s = pd.Series(['a_b_c_e_f', 'c_d_e_f_g', np.nan, 'f_g_h_i_j_k'])
s1=s.str.partition('_')
df1=df['brand'].str.partition(r'/')
'a_b_c_e_f'.partition("_")
# 右侧
s.str.rpartition('_')
df1=df.社区等级.str.rpartition('：')

# lower() 全部小写\upper() 全部大写
df1=df['brand'].str.lower()
df1=df['brand'].str.upper()

# find() 从左边开始，查找给定字符串的所在位置
s.str.find('d')
df1 = df['brand'].str.find('oo')
df1 = df['brand'].str.rfind('oo')



# index() 查找给定字符串的位置，注意，如果不存在这个字符串，那么会报错！
s.str.index('_')
df1=df.iloc[:2,:]['shop'].str.index('旗')
ii=list(b_list['brand']).index(cate.split('-')[-1])
'sdasd'.index('a')

# capitalize() 首字符大写
s.str.capitalize()
df.会员名.str.capitalize()

# swapcase() 大小写互换
s.str.swapcase()

# isalnum() 是否全部是数字和字母组成
s = pd.Series(['a_b_c_e_f', 'c_d_e_f_g', np.nan, 'f_g_h_i_j_k', '1234'])
s.str.isalnum()
df['href'].str.isalnum()

# isalpha() 是否全部是字母
s.str.isalpha()
df['href'].str.isalpha()

# isdigit() 是否全部都是数字
s.str.isdigit()
df['skuId'].str.isdigit()


# isspace() 是否空格
s.str.isspace()

# islower() 是否全部小写\isupper() 是否全部大写\istitle() 是否只有首字母为大写，其他字母为小写
s.str.islower()
s.str.istitle()

# numeric() 是否是数字
s.str.isnumeric()
df['price'].str.isnumeric()

# isdecimal() 是否全是数字
s.str.isdecimal()
df['price'].str.isdecimal()

# 调用index, 适用于dataFrame, series
df.index
list(df.index)








""" np_numpy """
# np.title
a = np.array([0,1,2])
np.tile(a,(2,3))

np.tile(a,(3,1))
np.tile(a,(3,5))

















""" [效率、代码简化、pythonic] """

## retuen - if
class HTTP:
    def get(self, url, return_json=True):
        r = requests.get(url)
        # （if-return） 作为一个特例  ，最后再 return

        if r.status_code !=200:
            return {} if return_json else ''
        return r.josn() if return_json else r.text


## 取反~
df1=df[~df['brand'].str.contains('好孩子')]
df1=df[~df['shop'].str.contains('(健|专|营|架)')]
fsegmentDataFrame = segmentDataFrame[
    ~segmentDataFrame.segment.isin(stopwords.stopword)
]

## 去重
a = [11,22,33,44,11,22]  
b = set(a)
b
set([33, 11, 44, 22])  

c = [i for i in b]
c  
[33, 11, 44, 22]  
df.drop_duplicates(subset=['社区等级'], keep='first', inplace=True)
df1=df.drop_duplicates(subset=['社区等级'], keep='last', inplace=False)

## 10-12 把函数作为参数传递
import re 
s = 'A8C3721D86'

def convert(value):
    matched = value.group()
    if int(matched) >=6:
        return'9'
    else:
        return'0'

r =re.sub('\d', convert, s)
print(r)


##  format{} dict, args, kwargs
def func(a,*args,**kwargs):
    print(a,args,kwargs)
    for k,v in kwargs.items():
        print('{}--{}'.format(k,v))

func(10,20,30,bj=30,sh=98)


def func(*args):
    for s in args:
        print(s)
        print('-'*30)

func(*['Q31','Q12','Q34FA'])






















"""  [加密解密] """
####    1. base64    (可逆)
import base64
# 加密
encodestr = base64.b64encode(r"This is a key123,@#$%田蕴川".encode(encoding='utf-8'))
print(encodestr) # 注意encodestr类型是byte,不是str
print(encodestr.decode())
type(encodestr)
len(encodestr)
s
# 解码
decodestr = base64.b64decode(encodestr)
print(decodestr.decode())

str_encode = base64.b64encode(r'天下有雪,江山如画,あいおえお,シャトル'.encode(encoding='utf-8'))
base64.b64decode(str_encode).decode()

####    2. pycrypto模块
#   https://www.dlitz.net/software/pycrypto/api/2.6/
import hashlib

hash = hashlib.sha256()
hash.update('admin'.encode('utf-8'))
print(hash.hexdigest())
len(hash.hexdigest())


#md5进行数据的加密
import hashlib
#md5加密
def encryption_md5(name):
    m = hashlib.md5()  #创建一个hashlib.md5()对象
    m.update(name.encode("utf8"))    #将参数转换为UTF8编码
    print(m.hexdigest())            #用十六进制输出加密后的数据
    print(len(m.hexdigest()))          #用十六进制输出加密后的数据
 
encryption_md5("lucy")
encryption_md5("hello world")
encryption_md5("luboyan")
encryption_md5("田蕴川")

import hashlib
import binascii
data = "this is a md5 test."
str_md5 = hashlib.md5("data".encode(encoding='UTF-8')).hexdigest()
len(str_md5)
print('MD5加密后为 ：' + str_md5)


string = 'You are the best! Fighting!!!'
string1 = b'You are the best! Fighting!!!'
 
string2 = bytes(string, encoding='utf-8')
string3 = string.encode('utf-8')
 
string_int1 = int(binascii.hexlify(string1), 16)
string_int2 = int(binascii.b2a_hex(string2), 16)
print(string_int1)
print(string_int2)


## 生成密码
# import string, random
source = list(string.ascii_letters)
source.extend([str(s) for s in range(0,10)])
captcha = random.sample(source, 6)

import shortuuid
shortuuid.uuid

""" [文件读取保存] """
import os
os.getcwd()

####    1. pickle
## to_pickle
picklepath=r'C:\Users\tianyunchuan\iCloudDrive\Desktop\_pyCodeSpyder_\_utils_\_dict\country.pickle'
df.to_pickle(picklepath)
df.to_pickle(r'./city.pickle')

## read_pickle
df = pd.read_pickle(r'C:\Users\tianyunchuan\iCloudDrive\Desktop\_pyCodeSpyder_\_utils_\_dict\city.pickle')
#picklepath='C:\\PY\\sunstar\\pickle\\'+'id'+df.itemid[0]+'_'+contentcount+'.pickle'
#df.to_pickle(picklepath)

####    2. [excel]读取、保存
import pandas as pd
picklepath=r'C:\Users\tianyunchuan\iCloudDrive\Desktop\_pyCodeSpyder_\_utils_\_dict\'
df.to_excel('{}country.xlsx'.format(picklepath),sheet_name='country') 
df = pd.read_excel('{}country.xlsx'.format(picklepath),sheet_name='country')
"""
DataFrame.to_excel(excel_writer, sheet_name='Sheet1', na_rep='',
 float_format=None, columns=None, header=True, index=True, 
 index_label=None, startrow=0, startcol=0, engine=None, 
 merge_cells=True, encoding=None, inf_rep='inf', verbose=True, 
 freeze_panes=None)
"""


""" 创建空表 """
import xlwt
work_book=xlwt.Workbook(encoding='utf-8')
sheet=work_book.add_sheet('sheet表名')
#sheet.write(0,0,'第一行第一列')
#sheet.write(0,1,'第一行第二列')
work_book.save('Excel表.xlsx')


def read_excel(io, sheet_name=0, header=0, skiprows=None,skip_footer=0,
index_col=None, names=None, usecols=None, parse_dates=False,
date_parser=None, na_values=None, thousands=None,
convert_float=True, converters=None, dtype=None,
true_values=None, false_values=None, engine=None,
squeeze=False, **kwds)


## 导入excel设置第一列为index
df = pd.read_excel(r'C:\Users\tianyunchuan\iCloudDrive\Desktop\_pyCodeSpyder_\_data\data_survey_proc\ca_mca_nabe.xlsx', index_col=0)


## 保存到不同工作页
# http://www.52codes.net/develop/shell/58896.html
from openpyxl import load_workbook 

#先创建一个 2017.xlsx 文件
book = load_workbook("2017.xlsx")
writer = pd.ExcelWriter("2017.xlsx", engine='openpyxl')
writer.book = book
df1.to_excel(writer,sheet_name="A")
df2.to_excel(writer,sheet_name="B")
writer.save()


####    3. 文件读取.[txt]
f = open(r'C:\temp/1.txt')
f.read()

with open(r'C:\temp/1.txt',"r",encoding='utf-8') as f:
    str = f.read()
nickname,skutype,ratecontent,ratedate = [],[],[],[]
nickname.extend(re.findall(r'displayUserNick":"(.*?)","structuredRateList',str))
skutype.extend(re.findall(re.compile(r'auctionSku":"(.*?)","anony'),str))
ratecontent.extend(re.findall(re.compile(r'rateContent":"(.*?)","fromMall'),str))
ratedate.extend(re.findall(re.compile(r'rateDate":"(.*?)","rateContent'),str))

#### 文件读取写入 [open]
fh = open(r'C:\Users\tianyunchuan\iCloudDrive\Desktop\_win_Trans_mac_\_dell_proc\00._EC_HHD\_tian_test\config_step1.yml', 'w', encoding='utf-8')
fh.write(part_1)
[fh.write('-{}\n'.format(s)) for s in list_raw_columns]
fh.close()

with open(r'C:\Users\tianyunchuan\iCloudDrive\Desktop\_win_Trans_mac_\_dell_proc\00._EC_HHD\_tian_test\config_step1.yml',"w",encoding='utf-8') as fh:
    fh.write(part_1)    
    [fh.write('-{}\n'.format(dict_columns.get(s, s))) for s in list_raw_columns]
    fh.write('\ndelete_column_name:\n')
    [fh.write('-{}\n'.format(s)) for s in list_complement]

# 'a' 表示数据追加
with open('login_log.txt', 'a') as fp:
    fp.write(log_line)


####    4. [csv]
## 文件读取
import csv, pandas as pd

def read_csv_reader():
    with open ('stock.csv', 'r') as fp:
        # reader 是一个迭代器
        reader = csv.reader(fp)
        next(reader)    # 不读取表头
        for x in reader:
            name = x[3]
            volumn = x[-1]
            print({"name":name,"volumn":volumn})

def read_csv_DictReader():
    with open('stock.csv','r') as fp:
        # 使用DictReader创建的reader对象
        # 不会包含表头文件
        # reader是一个迭代器，遍历这个迭代器，返回的是一个字典
        reader = csv.DictReader(fp)
        for x in reader:
            info = {'secShortName':x['secShortName'],'turnoverVol':x['turnoverVol']}
            print(info)

def read_csv_DataFrame():
    df = pd.read_csv (r'stock.csv')

if __name__ == '__main__':
    # read_csv_DictReader()
    read_csv_DataFrame()


## 处理中文乱码
raw_sa.to_csv(r'{}/data_proc/raw_sa_{}.csv'.format(PATH_HEAD,name_survey),index=False,encoding="utf_8_sig")

## 文件写入 
import csv

def write_csv_demo1():
    headers = ['username', 'age', 'height']
    values = [
        ('張三', 18, 180),
        ('李四', 19, 190),
        ('王五', 20, 160)
    ]

    with open('classroom.csv', 'w', encoding='utf-8', newline='') as fp:
        writer = csv.writer(fp)
        writer.writerow(headers)
        writer.writerows(values)


def write_csv_demo2():
    headers = ['username', 'age', 'height']
    values = [
        {'username':'张三','age':18,'height':180},
        {'username':'李四','age':19,'height':190},
        {'username':'王五','age':20,'height':160}
    ]
    with open('classroo1.csv','w',encoding='utf-8',newline='') as fp:
        writer = csv.DictWriter(fp,headers)
        # 写入表头数据的时候，需要调用writeheader方法
        writer.writeheader()
        writer.writerows(values)


if __name__ == '__main__':
    write_csv_demo2()    

## 指定列的格式 str
d_colToStr = {'订单编号':str,'狮王的物料编码':str,'BOM的组合编码':str,}
raw_1 = pd.read_csv(r'{}\鲲驰海外旗舰店2019年.csv'.format(PATH),dtype=d_colToStr)



####    5. pd.read_json 待添加

    
    


####    6. pd.read_hdf
hdf = pd.HDFStore(r'C:\_data\python_base_163_zjd\data\hdf.h5')
hdf['df'] = df
hdf_read = pd.HDFStore(r'C:\_data\python_base_163_zjd\data\hdf.h5')


""" [JSON] """
##    定义
# 1.轻量级的数据交换格式
# 2.强调一万遍，JSON是一种数据格式
# 3.字符串是JASON的表现形式、是JSON的载体
# .. JSON vs XML (XML使用越来越少)
# 4.易于阅读、易于解析、网络传输效率高、跨语言交换数据(！)
# 5. 应用场景（网站后台>浏览器）
# 6. JASON >> object, array (两种形式)

## json > python 类型
import json
json_to_python = {
        'object':'dict',
        'array':'list',
        'string':'str',
        'number_1':'int',
        'number_2':'float',
        'true':'True',
        'false':'False',
        'null':'None'
        }


### html json string to dict
  r = requests.get(s1, headers=HEADERS)
  html = r.text
  json.dumps(html)

###    10-17 反序列化 (from Json)
## 反序列化定义：从字符串转换到某种数据结构的过程
import json
json_str = '{"name":"qiyue","age":18}'
stu = json.loads(json_str)

stu.get('name')
stu['name']

# array
json_str = '[{"name":"qiyue","age":18,"flag":false},{"name":"Dec","age":20}]'
stu = json.loads(json_str)

import pandas as pd
df = pd.DataFrame(stu)

## json转化二进制
import json
text={"status":1,"content":{"from":"zh","to":"ja","vendor":"ciba","out":"\u4eac\u6771\u304c\u81ea\u55b6\u3067\u8cfc\u5165\u3057\u305f\u3082\u306e\u3067\u3001 \u30ad\u30e3\u30f3\u30da\u30fc\u30f3\u4fa1\u683c\u306f\u304a\u5f97\u3067\u3001","ciba_out":"","err_no":0}}
json.loads(response.text)

###    10-18 序列化 (to JSON)
stu = [
       {"name":"qiyue","age":18,"flag":False},
       {"name":"Dec","age":20}
      ]
## dict > json字符串
import json
json_str = json.dumps(stu)
type(json_str)

json.dumps(books, default=lambda  o:o.__dict__)
var = json.dumps(json_to_python, default=lambda  o:o.__dict__)


""" [ip代理 代理服务器] """
import requests
s = requests.session()
s.proxies = {'http': '106.14.241.155:80'}
s.get(r'http://www.dianping.com/shop/92164031/review_more?pageno=1',headers=headers)
content=requests.get(url_base,headers=headers).text


""" [函数] [参数] """s
####        10-12 把函数作为参数传递
import re 
s = 'A8C3721D86'

def convert(value):
    matched = value.group()
    if int(matched) >=6:
        return'9'
    else:
        return'tian'

r =re.sub('\d', convert, s)

print(r)


####        8-3 如何让函数返回多个结果
def add(x,y):
    result = x + y
    return result

 a =add(100,99)


def damage(skill1,skill2):
    return skill1*2+100,skill2*100000+345

a,b = damage(100,200)

damage_skill_1,damage_skill_2 = damage(100,200)
r = damage(23,45)


####    8-4 序列解包与链式赋值
a,b,c = 1,2,3
z = 1,2,3,4
a,b,c,d = z

x = [1, 2, 3]
y = [4, 5, 6]
z = [7, 8, 9]
u = zip(x,y,z)
type(u)
a,b,c = zip(*u)


####    8-6 默认参数 （必传参数在前，默认参数都在后面！）

def print_files(name,gender='',age=18,college='幸福村'):
    print(f'我叫{name}')
    print(f'我今年{age}岁')
    print(f'我是{gender}生')
    print(f'我在{college}上学')

print_files('mike','男',200,'新福村')
print_files('mike')
print_files('jonny',age=20)


####    8-7 可变参数
from functools import reduce
def demo(*param):
    print(param)
    print(type(param))
    print(reduce(lambda x,y:x+y,param))
demo(1,2,3,4,5,6)
lst = [1,2,3,4,5,6,7,8,9,10]
demo(*lst)

def demoTuple(*param):       # 传入元祖
    print(param)
    print(type(param))
a = (1,2,3,4,5)
demoTuple(*a)    # *作用是解包
demoTuple(a)     # 不是元祖，是

def demoMulti(param1,param2=2,*param):       # 多种参数的顺序
    print(param1)
    print(param2)
    print(param)

demoMulti('a',2,1,2,3)


def demoMulti(param1,*param,param2=2):       # 多种参数的顺序
    print(param1)
    print(param)
    print(param2)

demoMulti('a',1,2,3,'param2')


def demo(*param):
    for i, element in enumerate(param):
        print(i,element)
demo(1,2,3,4,5,)

def demo(*param):
    for element in param:
        print(element)
demo(1,2,3,4,5)


####    8-8 关键字可变参数
def squsum(*param):
    sum1 = 0
    for i in param:
        sum1 += i*i
    print(sum1)

squsum(1,2,3,4,5,6)
squsum(*[1,2,3,4,5,6])

# 任意个数的关键字参数 (可以不传参)
def city_temp(**param):
    print(param)
    print(type(param))
    for k,v in param.items():
        print(f'{k}:{v}')
        

city_temp(bj='32c',sh='23c',gz='23c')

a = {'sh':34,'bj':23,'dsaf':'tian'}
city_temp(**a)





""" [统计] [stat] """
from pandas import DataFrame
df = DataFrame({'key1':['a','a','b','b','a'],'key2':['one','two','one','two','one'],'data1':np.random.randn(5),'data2':np.random.randn(5)})
#[Out]#       data1     data2 key1 key2
#[Out]# 0  0.439801  1.582861    a  one
#[Out]# 1 -1.388267 -0.603653    a  two
#[Out]# 2 -0.514400 -0.826736    b  one
#[Out]# 3 -1.487224 -0.192404    b  two
#[Out]# 4  2.169966  0.074715    a  one

df.describe()
df.describe(include='all')


""" 中继点 """
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
#%matplotlib inline
%matplotlib qt     #独立图显示


""" 1.1  mean / median """
mtcars=pd.read_csv(r'C:\_TianProc\_data_\data_class\Python_basic_class_163_zjd_data/mtcars.csv')
mtcars.head()
























''' [groupby] '''
####    1. 分组groupby 
# https://blog.csdn.net/Asher117/article/details/85614034

import pandas as pd
""" 1.GroupBy过程 """
##    首先看看分组聚合的过程，主要包括拆分（split）、应用（Apply）和合并（Combine）

""" 2.创建数据 """
ipl_data = {'Team': ['Riders', 'Riders', 'Devils', 'Devils', 'Kings',
         'kings', 'Kings', 'Kings', 'Riders', 'Royals', 'Royals', 'Riders'],
         'Rank': [1, 2, 2, 3, 3,4 ,1 ,1,2 , 4,1,2],
         'Year': [2014,2015,2014,2015,2014,2015,2016,2017,2016,2014,2015,2017],
         'Points':[876,789,863,673,741,812,756,788,694,701,804,690]}
df = pd.DataFrame(ipl_data)


""" 3.将df按照Team列分组 """
##    分组之后的grouped是个DataFrameGroupBy对象
grouped = df.groupby('Team')
grouped

""" 4.查看df按照Team分组之后结果 """
##  按照不同的Team值来分组
grouped.groups

""" 5.分组后的结果对Points求和 """
grouped['Points'].sum()
grouped['Points'].mean()
grouped['Points'].count()   # 计数

# count, mean, max, min, count, mediam, std, var, prod(乘积), first, last

result = data.groupby(by=['手机品牌'],as_index=False)['月消费（元）'].agg({'月消费': numpy.sum})  # index=

""" 6.根据任意长度适当的数组分组 """
import numpy as np
key1 = np.array(list('abababbaaabb'))
df.groupby(key1).Team.count()

""" 7.对分组进行迭代 """
## GroupBy分组产生的是一组二元元组，有分组名和数据块组成。即(分组名、数据块)。
for name,group in df.groupby('Team'):
    print(name)
    print(group)
    print('*******分隔符*********')

## 另外，对于多重建分组的情况，元组的第一个元素将是由元组组成。
## 即（(分组名1,分组名2)、数据块）。
for (name1,name2),group in df.groupby(['Team','Rank']):
    print(name1)
    print(name2)
    print(group)
    print('*******分隔符*********')
    
""" 8.在不同轴上分组 """
## GroupBy默认是在axis=0轴上进行分组的，也可以在axis=1轴上进行分组聚合，不过用的相对较少。
df.dtypes
grouped = df.groupby(df.dtypes, axis=1)
grouped.groups

""" 9.通过字典或Series进行分组 """
people = pd.DataFrame(np.random.randn(5, 5),
                      columns=['a', 'b', 'c', 'd', 'e'],
                      index=['Joe', 'Steve', 'Wes', 'Jim','Travis'])
people.iloc[2:3, [1, 2]] = np.nan
people
## 根据字典聚合
mapping = {'a': 'red', 'b': 'red', 'c': 'blue',
           'd': 'blue', 'e': 'red', 'f' : 'orange'}
by_column = people.groupby(mapping, axis=1)
by_column.sum()
## 根据Series聚合
map_series = pd.Series(mapping)
people.groupby(map_series, axis=1).count()

""" 10.通过函数进行分组 """
z = people.groupby(len).sum()


""" 11.函数、数组、列表、字典、Series组合分组 """
key_list = ['one', 'one', 'one', 'two', 'two']
people.groupby([len, key_list]).min()

""" 12.根据索引级别分组 """
## 回到最初的DataFrame，给他重新定义成双层索引，并且给索引命名
df.columns = ([['a','a','a','b'],['Team', 'Rank', 'Year', 'Points']])
df.columns.names = ['one','two']
df
## 之后对索引名为one的进行分组聚合
df.groupby(level='one',axis=1).count()

""" 12.1 多函数聚合 """
# 其中多函数聚合中也可以使用自定义函数。
df.columns = ['Team','Rank','Year','Points']
df.groupby('Team')['Points'].agg(['sum','mean','std','count'])


""" 13.apply：一般性的“拆分-应用-合并” """
## 定义函数：
def top(df,n=2,column='Points'):
    return df.sort_index(by=column,ascending=False)[:n]
## 应用
df.groupby('Team').apply(top)

## 同时给apply函数传入参数：
df.groupby('Team').apply(top,n=3)


import pandas as pd

df = pd.read_excel(r'C:\Users\tianyunchuan\iCloudDrive\_data_\data_training\training_01.xlsx', keep_default_na=False)



####    [crosstab]  （适合分类变量）
#crosstab(index, columns, values=None, rownames=None, colnames=None, aggfunc=None,
#margins=False, dropna=True, normalize=False)
temp = pd.crosstab(df['city_cate'], df['Q1'])
temp = pd.crosstab(df['city_cate'], df['Q2'])

temp = pd.crosstab([df['city_cate'],df['gender_cate']], df['Q2'])
z  =pd.crosstab(df['q1'], df['num_age'])


temp = pd.crosstab(df['brand'],df['shop'])                                              # cross
temp = pd.crosstab([df['shop'],df['timer_1']],df['brand'])                              # 表侧两重
temp = pd.crosstab(df['brand'],[df['timer_1'],df['shop']])                              # 表头两重
temp = pd.crosstab(df['city_cate'],df['Q2'],margins=True,margins_name='Total')            # 加合计
temp = pd.crosstab(df['city_cate'],df['Q2'],margins=True,normalize=True)                # 计算百分比
pd.crosstab(df['city_cate'],df['Q3_1'],margins=True,normalize=True)                # 计算百分比


####    [pivot_table]  （适合数值型变量）
agg = pd.pivot_table(df,index=['shop'])                                                 # 默认均值，df内所有数值变量
agg = pd.pivot_table(df,index=['shop','timer_1'])                                       # 双重，所有数值变量
agg = pd.pivot_table(df,index=['shop'],values='monetary')                               # 某变量
agg = pd.pivot_table(df,index=['shop'],values='monetary',aggfunc=np.sum,margins=True)
agg = pd.pivot_table(df,index=['shop'],values='monetary',aggfunc=np.mean)

agg = pd.pivot_table(df,index=['brand'],columns=['timer_1'],values='monetary',aggfunc=np.sum,margins=True)



####    [pandas] [pd]
# https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.factorize.html
## factorize  将对象编码为枚举类型或分类变量。pd.factorize(['b', 'b', 'a', 'c', 'b'])
codes, uniques = 
codes, uniques = pd.factorize(['b', 'b', 'a', 'c', 'b'], sort=True)
codes, uniques = pd.factorize(['b', None, 'a', 'c', 'b'])

# codes   -> [0 0 1 2 0]
# uniques -> ['b' 'a' 'c']

##
cat = pd.Categorical(['a', 'a', 'c'], categories=['a', 'b', 'c'])
codes, uniques = pd.factorize(cat)


## 读取指定行数
df = pd.read_csv(r'iris.csv', nrows=50)


## 获取、查看列的dtype
df = pd.read_csv(r'C:\_TianProc\_data_\data_class\Python_basic_class_163_zjd_data\titanic.csv')
df.select_dtypes(include=['object', 'int'])
df.select_dtypes(exclude=['float64'])
help(df.select_dtypes)









""" [other] """
##  warning
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings("ignore", category=RuntimeWarning)


##  iter
x = iter(["apple", "banana", "cherry"])
print(next(x))
print(next(x))
print(next(x))

l=[2,3,4]
iterl = iter(l)
next(iterl)
next(iterl)
next(iterl)


## 整数除法
X / 2
X // 2 (整数除法)


""" [selenium] """
### chromedriver的绝对路径
driver_path = var.driver_path_firefox
### 初始化一个driver，并且指定chromedriver的路径
driver = webdriver.Firefox(executable_path=driver_path)
""" 账号登录 """
driver.get('https://passport.weibo.cn/signin/login')
time.sleep(6)
# 通过page_source获取网页源代码
print(driver.page_source)
# 填写用户名、密码登录
#driver.find_element_by_id ("loginName").send_keys ("13916******")
#driver.find_element_by_id ("loginPassword").send_keys ("tt****20_01")
driver.find_element_by_xpath ('//*[@id="loginName"]').send_keys ("13916******")
driver.find_element_by_xpath ('//*[@id="loginPassword"]').send_keys ("tt7")
driver.find_element_by_id ('loginAction').click()

### chrome
## chrome.exe 驱动下载 http://selenium.googlecode.com/svn/trunk/docs/api/py/index.html
from selenium import webdriver
option = webdriver.ChromeOptions()
option.add_experimental_option('excludeSwitches', ['enable-automation']) #这里去掉window.navigator.webdriver的特性

domain = "https://www.taobao.com/"
# domain = "http://news.cnblogs.com/"


#下面的chromedriver.exe使用特殊的可执行文件，去掉了$cdc_lasutopfhvcZLmcfl等特性
browser = webdriver.Chrome(executable_path=r"C:\Users\tianyunchuan\pySurvey\_utils_\chromedriver_win32\chromedriver.exe", options=option)
browser.get(domain)
browser.page_source

## headless 浏览器无头
PATH_DRIVER = r'C:\Users\tianyunchuan\iCloudDrive\_spyder_\_utils_\webDriver\chromedriver.exe'
from selenium.webdriver import Chrome, ChromeOptions
option = ChromeOptions()
option.add_argument('--headless') #隐藏浏览器
option.add_argument('--no-sandbox') #隐藏浏览器
browser = Chrome(executable_path=PATH_DRIVER, options=option)


## frame 登录
import time
from selenium import webdriver
# 创建浏览器对象
browser = webdriver.Firefox(executable_path = r'C:\_TianProc\_spyder_\_utils_\webDriver/geckodriver.exe')

# 打开163邮箱网页
browser.get("https://mail.163.com/")
# 这里睡眠5秒,因为网页登录加载很慢,直接获取元素获取不到.
time.sleep(5)
# 通过switch_to.frame 进入到frame富文本框里面
browser.switch_to.frame(browser.find_element_by_xpath('//iframe[@scrolling="no"]'))
browser.find_element_by_name("email").send_keys("13916******")
browser.find_element_by_name("password").send_keys("Tt******")
browser.find_element_by_id("dologin").click()

## 获取多个同名的class
  all_more = driver.find_elements_by_class_name("more-words") 
  for i in all_more:
      i.click()
      time.sleep(0.2)

##
switch_to_window方法已过期，使用switch_to.window方法来代替


""" [xlml] [xpath] """
####    读取本地 txt文件为html
with open(r'C:\Users\tianyunchuan\Desktop\arena\13.txt',"r",encoding='utf-8') as f:
    text = f.read()
html = etree.HTML(text) 

## 查看字符集
print(r.encoding)
print(r.apparent_encoding)

## 转换字符集
text = resp.content.decode('gbk')

## 使用contains函数:
//div[contains(@id,'test')]

""" [requests] """
## close
## https://www.cnblogs.com/prac/p/whyDontUseResponseClose.html



""" [re] """
# -*- coding: utf-8 -*-
# https://www.cnblogs.com/austinjoe/p/9492790.html

import re

source = '<span id="reviewCount" class="item">田蕴川条评论 </span>'
reviewCount = re.findall(r'<span id="reviewCount" class="item">(.*?)条评论 </span>',source, re.DOTALL)

s = '基础知识概要：1-1 http，url'

""" 1. start """

re.findall('\d', s)


""" 2.普通字符、元字符 
普通字符：'C|C++|C#|Python|Javascript'
元字符：‘\d’
"""


""" 3. 字符集 """
s = 'aac,abc,acc,adc,aec,afc,ahc,abg,pqr'

# 任意字符
re.findall('a[ac]c', s)

import re
s='这个包装看上去时髦'
words=['时髦','流行','时尚']
pat=re.compile('|'.join(words))
1 if pat.search(s) else 0 

# 找到所有非数字
add2 = '沪路1142号佑米造型qw234'
re.findall('\D', add2)

# 取反
re.findall('a[^z]c', s)

# 区间
re.findall('a[a-c]c', s)
re.findall('a[^a-c]c', s)


""" 4.概括字符集
'\d'    >>>>    数字
'\D'    >>>>    英文字母
'\s'    >>>>    空白字符
.       >>>>    换行符（\n）之外其他所有字符
"""
s = 'abc123ik9080 -0-0-11-2-3 -4-'
re.findall('\d', s)
re.findall('[6-9]', s)
re.findall('[^6-9]', s)
re.findall('\D', s)
re.findall('\s', s)
re.findall('.', s)

""" 5.数量词 """
s = 'abc123ik9080 -0-0-11-abc2-3 ab-4-'
re.findall('[a-c][a-c]', s)
re.findall('[a-c]{3}', s)
re.findall('[a-c]{1,3}', s)


""" 6. 贪婪与非贪婪 (默认是贪婪的） """
s = 'abc123ik9080 -0-0-11-abc2-3 ab-4-'
re.findall('[a-c]{1,3}?', s)

text = "0123456"
ret = re.findall('\d+?',text)
ret = re.findall('\d+',text)


""" 7. 匹配次数 ???????? """  
# 匹配字符 0-无数次    >>  * 
s = 'C|C++|C#|Python|Javascript|Python|python|PythonPythonPython'
re.findall('Python*', s)

# 匹配字符 1-无数次    >>  +
re.findall('Python+', s)

# 匹配字符 0-1次      >>  ?
re.findall('Python?', s)

# 非贪婪应用场景
re.findall('Python{1,2}', s)
re.findall('Python{1,2}?', s)
re.findall('Python{1,2}*', s) 


""" 8. 边界匹配符 ????????"""
s = '1000423424000'
re.findall('^000', s)
re.findall('^1000', s)
re.findall('000$', s)

# 以**开头的
text = "13916380622"
#ret = re.match('^139+\d',text)
ret = re.findall('^139+\d+',text)

# 以**结尾的
text = "xxx@163.com"
ret = re.match('\w+@163.com$',text)


""" 9. 组 """
s = 'PythonPythonPythonJSPythonPythonPytn'
re.findall('(Python){3}(JS)', s)
re.findall('(Python){2}(JS)', s)
re.findall('(Python)(JS)', s)

# # 注意： [abc] 中括号是或关系，(abc)是且关系
re.findall('["Python","JS"]', s)


""" 10.匹配模式参数 ???? """
s = 'C#PythonPythonPythonJSPythonPythonPytn'
# re.I 忽略大小写
re.findall('c#', s)
re.findall('c#', s, re.I)
           
# 
re.findall('c#.{1}', s,re.I)
re.findall('c#.{2}', s,re.I)
re.findall('c#.{1}', s, re.I|re.S)

# re.I  使匹配对大小写不敏感
# re.L  做本地化识别（locale-aware）匹配
# re.M  多行匹配，影响 ^ 和 $
# re.S  使 . 匹配包括换行在内的所有字符
# re.U  根据Unicode字符集解析字符。这个标志影响 \w, \W, \b, \B.
# re.X  该标志通过给予你更灵活的格式以便你将正则表达式写得更易于理解。



""" 11.re.sub 替换 """
# pattern : 正则中的模式字符串。
# repl : 替换的字符串，也可为一个函数。
# string : 要被查找替换的原始字符串。
# count : 模式匹配后替换的最大次数，默认 0 表示替换所有的匹配。
# flags : 编译时用的匹配模式，数字形式。
# 前三个为必选参数，后两个为可选参数。

s = 'PythonC#\nJavaPHPC#'
re.sub('C#','GO', s)
re.sub('C#','GO', s, count=1)
# replace函数 lang.replace('C#','GO')

#https://www.runoob.com/python3/python3-reg-expressions.html
import re
phone = "2004-959-559 # 这是一个电话号码"
 
# 删除注释
num = re.sub(r'#.*$', "", phone)
print ("电话号码 : ", num)
 
# 移除非数字的内容
num = re.sub(r'\D', "", phone)
print ("电话号码 : ", num)

# repl 参数是一个函数
# 将匹配的数字乘于 2
def double(matched):
    value = int(matched.group('value'))
    return str(value * 2)
 
s = 'A23G4HFD567'
print(re.sub('(?P<value>\d+)', double, s))


## 把函数当成参数传递 例1
def convert(value):
    matched = value.group()
    return f'!!!{matched}!!!'

## 把函数当成参数传递 例2
re.sub('C#',convert, s, count=0)  

s = 'A8C3721D86'

def convert(value):
    matched = value.group()
    if int(matched) >=6:
        return'9'
    else:
        return'0'

r =re.sub('\d', convert, s)
print(r)
# 题外
s = '北京上海广州深圳' 
def if_condition(value):
    if value in '北京上海广州':
        return 'Tier_1'
    else:
        return 'Tier_other'
if_condition('广州')

import re
def convert(value):
    matched = value.group()
    return f'!!!{matched}!!!'
re.sub('上海', if_condition, s, count=0)   

s_li = ['上海','北京','广州','成都']   
'上海' in s_li

# 同时替换多个
a='1991年10月12日'
a = re.sub(r'年|月|日', "-",a.strip('日'))




""" 13. search march"""
s = '8C3721D86'
r = re.match('\d',s)
r.group()
r.span()
r = re.findall('\d',s)

""" 14. 分组 """
s = 'life is short, i use python, ilove python'
r = re.search('life(.*)python(.x)python',s)
r.group(0)

""" re多条件匹配 """
re.findall(r'[\u4e00-\u9fa5]+|\d+', "111")
re.findall(r'[\u4e00-\u9fa5]+|\d+', "田大发")

reObj = re.compile(r'[\u4e00-\u9fa5]+|\d+|\(|\)|\:|\,|\?|\*|\&|\^|\%|\#|\-')
[re.findall(reObj, s) for s in l_addr]


""" 99. 抽取 """
s = 'use我们是方法为we'
re.findall(re.compile(r'use(.*?)we'), s)

last_page = re.findall(re.compile(r'共(.*?)页'),'共10页')


# 20. |：匹配多个字符串或者表达式：
text = "!!https"
ret = re.findall('(ftp|http|https)$',text)


# 实例
text = "apple price is $299"
ret = re.search("\$\d+",text)
ret = re.findall("\$\d+",text)[0]


text = '\n'
print(text)

text = "apple's price $99,orange's price is $10"
ret = re.search('.*(\$\d+).*(\$\d+)',text)
print(ret.group(0))
print(ret.group(1))
print(ret.group(2))
print(ret.group(1,2))
print(ret.groups())


""" [picture] [图片] """
from PIL import Image
import matplotlib.pyplot as plt
img=Image.open(r'C:\Users\tianyunchuan\iCloudDrive\_data_\data_pic\earth.jpg')
plt.imshow(img)
plt.show()

###
import matplotlib.pyplot as plt
from matplotlib.image import imread
 
img = imread(r'/kaggle/input/picture/p003.jpg')
plt.imshow(img)
 
plt.show()


###

train = pd.read_csv('/kaggle/input/mnist-in-csv/mnist_train.csv')
train.head(2)

train.iloc[:1,1:]
np.array(train.iloc[:1,1:])

import matplotlib.pyplot as plt
from matplotlib.image import imread

plt.imshow(np.array(train.iloc[:1,1:]).reshape(28,28),cmap='bone')
 
plt.show()




###
import cv2

img = cv2.imread(r'C:\Users\tianyunchuan\iCloudDrive\_data_\data_pic\earth.jpg',cv2.IMREAD_GRAYSCALE)
cv2.imshow('img', cv2.resize(img,(200, 200)))
cv2.imshow('image',img)
cv2.waitKey(0)
cv2.destroyAllWindows()


""" [onehot] """
## bySelf
browser = ["Firefox", "Chrome", "Safari", "Internet Explorer"]
gender = ["male", "female"]
 
def onehotEncoding(instance1, instance2, class1, class2):
    temp1, temp2 = [0] * len(class1), [0] * len(class2)
    temp1[class1.index(instance1)] = 1
    temp2[class2.index(instance2)] = 1
    return temp1 + temp2
 
for i in browser:
    for j in gender:
        print(onehotEncoding(i, j, browser, gender))


## by skLearn
from numpy import array
from numpy import argmax
from sklearn.preprocessing import LabelEncoder
from sklearn.preprocessing import OneHotEncoder
# define example
data = ['cold', 'cold', 'warm', 'cold', 'hot', 'hot', 'warm', 'cold', 'warm', 'hot']
values = array(data)
print(values)
# integer encode
label_encoder = LabelEncoder()
integer_encoded = label_encoder.fit_transform(values)
print(integer_encoded)
# binary encode
onehot_encoder = OneHotEncoder(sparse=False)
integer_encoded = integer_encoded.reshape(len(integer_encoded), 1)
onehot_encoded = onehot_encoder.fit_transform(integer_encoded)
print(onehot_encoded)
# invert first example
inverted = label_encoder.inverse_transform([argmax(onehot_encoded[0, :])])
print(inverted)

## by pandas.get_dummies
pandas.get_dummies(data, prefix=None, prefix_sep='_', dummy_na=False, columns=None, sparse=False, drop_first=False)[source]

import pandas as pd
df = pd.DataFrame([  
            ['green' , 'A'],   
            ['red'   , 'B'],   
            ['blue'  , 'A']])  

df.columns = ['color',  'class'] 
pd.get_dummies(df) 

pd.get_dummies(df)
pd.get_dummies(df.color)
df = df.join(pd.get_dummies(df.color))


needcode_cat_columns = ["Pclass", "Sex", "SibSp", "Parch", "Embarked"]
df_coded = pd.get_dummies(
    df_train,
    columns=needcode_cat_columns,
    prefix=needcode_cat_columns,
    dummy_na=True,
    drop_first=True
)


""" color 颜色 """
from pysankey2.utils import setColorConf
colors = setColorConf(12,colors='Pastel1')


""" [crawler] """
import random
user_agent = [
    "Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50",
    "Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50",
    "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:38.0) Gecko/20100101 Firefox/38.0",
    "Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729; InfoPath.3; rv:11.0) like Gecko",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0)",
    "Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1)",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
    "Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
    "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11",
    "Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Maxthon 2.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; TencentTraveler 4.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; The World)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SE 2.X MetaSr 1.0; SE 2.X MetaSr 1.0; .NET CLR 2.0.50727; SE 2.X MetaSr 1.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; 360SE)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Avant Browser)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)",
    "Mozilla/5.0 (iPhone; U; CPU iPhone OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5",
    "Mozilla/5.0 (iPod; U; CPU iPhone OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5",
    "Mozilla/5.0 (iPad; U; CPU OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5",
    "Mozilla/5.0 (Linux; U; Android 2.3.7; en-us; Nexus One Build/FRF91) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1",
    "MQQBrowser/26 Mozilla/5.0 (Linux; U; Android 2.3.7; zh-cn; MB200 Build/GRJ22; CyanogenMod-7) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1",
    "Opera/9.80 (Android 2.3.4; Linux; Opera Mobi/build-1107180945; U; en-GB) Presto/2.8.149 Version/11.10",
    "Mozilla/5.0 (Linux; U; Android 3.0; en-us; Xoom Build/HRI39) AppleWebKit/534.13 (KHTML, like Gecko) Version/4.0 Safari/534.13",
    "Mozilla/5.0 (BlackBerry; U; BlackBerry 9800; en) AppleWebKit/534.1+ (KHTML, like Gecko) Version/6.0.0.337 Mobile Safari/534.1+",
    "Mozilla/5.0 (hp-tablet; Linux; hpwOS/3.0.0; U; en-US) AppleWebKit/534.6 (KHTML, like Gecko) wOSBrowser/233.70 Safari/534.6 TouchPad/1.0",
    "Mozilla/5.0 (SymbianOS/9.4; Series60/5.0 NokiaN97-1/20.0.019; Profile/MIDP-2.1 Configuration/CLDC-1.1) AppleWebKit/525 (KHTML, like Gecko) BrowserNG/7.1.18124",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows Phone OS 7.5; Trident/5.0; IEMobile/9.0; HTC; Titan)",
    "UCWEB7.0.2.37/28/999",
    "NOKIA5700/ UCWEB7.0.2.37/28/999",
    "Openwave/ UCWEB7.0.2.37/28/999",
    "Mozilla/4.0 (compatible; MSIE 6.0; ) Opera/UCWEB7.0.2.37/28/999",
]


headers = {'User-Agent': random.choice(user_agent),  # 随机选取头部代理,防止被屏蔽
           'Connection': "keep-alive",
           'Host': "s3plus.meituan.net",
           'referer': 'http://www.dianping.com/',
           'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
           }

response_woff = requests.get(url, headers=headers).content


""" [断点调试] """
import pbd;pdb.set_trace()

"""
#### PBD 常用指令（在（pdb) 后面输入并回车即可：
h --> 显示帮助，共有哪些指令 #-- help
h n --> 显示关于n指令的帮助 #-- help next
n --> 执行当前待执行的语句 #-- next
s --> 如果下一条语句是个函数，则此条指令跳入函数内部 #-- step 
r --> 跳入函数后想直接执行到该函数结束 #-- return
l --> 列出上下代码 #-- list
ll -->列出所有代码 #-- long list
c --> 继续执行，直到遇到断点 #-- continue
q --> 退出程序 #--quit
"""



[class]
""" 限制类属性 """
# -*- coding: utf-8 -*-
class Person:
    __slots__ = ['age']
    pass

p1 = Person()
p1.age = 18
p1.num = 2

# AttributeError: 'Person' object has no attribute 'num'

#### 实例方法、类方法、静态方法
class Person:
    def eat2(self):
        print("这是一个实例方法", self)

    @classmethod
    def leifangfa(cls):
        print('这是一个类方法', cls)
        
    @staticmethod
    def jingtaifangfa():
        print("这是一个静态方法")
        
        
p = Person()
p.eat2()

Person.leifangfa()
        
Person.jingtaifangfa()



""" [装饰器] [decorator]"""
## 无参数
def deco(func):
    def wrapper():
        print(time.time())
        func()
    return wrapper

@deco
def f1():
    print("This is func")

f1()

## 有参数
import time


def deco(func):
    def wrapper(*args, **kwargs):
        print(time.time())
        func(*args, **kwargs)

    return wrapper


@deco
def f1(func_name):
    print("This is func: {}".format(func_name))


@deco
def f2(func_name_1, func_name_2):
    print("This is func: {}, {}".format(func_name_1, func_name_2))


f1('test-func')

f2('train', 'test')


## 用于类中
class Calc:

    def deco_check_num(func):
        def wrapper(self, n):
            if not isinstance(n, int):
                raise TypeError('您输入数值的类型有问题')
            return func(self, n)
        return wrapper

    @deco_check_num
    def __init__(self, value):
        self.__result = value

    @deco_check_num
    def jia(self, n):
        self.__result += n

    @deco_check_num
    def jian(self, n):
        self.__result -= n

    def log(self):
        print('当前运算结果为：{}'.format(self.__result))

c = Calc(6)
c.jia(int('12'))
c.jian(1)
c.log()




""" [yield] """

def read_file(fpath): 
    BLOCK_SIZE = 1024 
    with open(fpath, 'rb') as f: 
        while True: 
            block = f.read(BLOCK_SIZE) 
            if block: 
                yield block 
            else: 
                return

""" [with] [上下文管理器] [context] """
import contextlib
@contextlib.contextmanager
def ze():
    try:
        yield
    except ZeroDivisionError as e:
        print('error', e)

x=1
y=0
with ze():
    x/y

## https://www.cnblogs.com/python1111/p/16541808.html
## 自定义上下文管理器来对软件系统中的资源进行管理，比如数据库连接、共享资源的访问控制等。

class With_work(object):
    def __enter__(self):
        """进入with语句的时候被调用"""
        print('enter called')
        return "xxt"

    def __exit__(self, exc_type, exc_val, exc_tb):
        """离开with的时候被with调用"""
        print('exit called')


with With_work() as f:
    print(f)
    print('hello with')

'''
enter called
xxt
hello with
exit called
'''


class Test:
    def t(self):
        print('tttt')

    def close(self):
        print ('资源释放')


import contextlib
with contextlib.closing(Test()) as t_obj:
    t_obj.t()


# with open(r'T恤_1.png', 'rb') as from_file:
#     with open('T恤_2.png', 'wb') as to_file:
#         from_contents = from_file.read()
#         to_file.write(from_contents)


with open(r'T恤_1.png', 'rb') as from_file, open('T恤_2.png', 'wb') as to_file:
    from_contents = from_file.read()
    to_file.write(from_contents)



#### 调用数据库
import pymysql
class con:
    def __init__(self, host, port, db, user, pwd):
 
        self.conn = pymysql.connect(
            host=host,   
            port=port,
            user=user,
            password=pwd,
            database=db,
            cursorclass=pymysql.cursors.DictCursor
            )
        
        self.cur = self.conn.cursor()
        print("连接已建立")
 
    def __str__(self):
 
        return "pymysql上下文管理器"
 
    def __enter__(self):
 
        # 返回类对象本身
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
 
        self.cur.close()
        self.conn.close()
        print("连接已关闭")
 
        # print(exc_type, "异常")
        # print(exc_val, "异常")
        # print(exc_tb, "异常")
    
    # 查询单条数据
    def read_one(self, sql):
 
        try:
            self.cur.execute(sql)
            res = self.cur.fetchone()
        
        except Exception as e:
 
            print(e)
        
        return res

    def read_all(self, sql): 
        try:
            self.cur.execute(sql)
            data = self.cur.fetchall()    
            cols = [col[0] for col in self.cur.description]
        except Exception as e: 
            print(e)        
        return data, cols

if __name__ == "__main__":
    with Get_data_by_pymysql(HOST, 3306, "world", "root", PASSWORD) as db: 
        print("这是一个", db) 
        result = db.read_all(" select * from `city` ") 
        df = pd.DataFrame(result[0], columns = result[1]) 


""" [语法糖] """

## 交换值
a = 2
b = 100
a, b = b, a

## 链接判断
a = 95
if 90 <= a < 100:
    print('优秀')

## 多位数字
a = 10_0000_0000
print(a)

## 列表拼接
a = [1, 2, 3]
b = [4, 5, 6]
print(a + b)

## 迭代对象拆分
x, y, z = [1, 2, 3]
print(x)

x, y, z = (1, 2, 3)
print(y)

# 反向打包
b = x,y,x



""" [np.array] """

#### 1. 基本操作
import numpy as np

np.__version__

nparr = np.array([i for i in range(10)])  # array([0, 1, 2, 3, 4, 5, 6, 7, 8, 9])

nparr[5] = 100

nparr.dtype  # dtype('int64')


#### 2. 其他创建np.array的方法
np.zeros(10)    #array([0., 0., 0., 0., 0., 0., 0., 0., 0., 0.])

np.zeros((3,2))
np.zeros(shape=(3,2), dtype=float)
"""
array([[0., 0.],
       [0., 0.],
       [0., 0.]])
"""

np.ones((5,2))

np.arange(0, 20, 2)   # array([ 0,  2,  4,  6,  8, 10, 12, 14, 16, 18])

np.linspace(0, 20, 10)
"""
array([ 0.        ,  2.22222222,  4.44444444,  6.66666667,  8.88888889,
       11.11111111, 13.33333333, 15.55555556, 17.77777778, 20.        ])
      """

np.random.randint(0, 100, 5)    # array([74, 94, 49, 82, 31])
np.random.randint(0, 100, size=5)

np.random.randint(10, size=(3,5))
"""
array([[3, 7, 3, 3, 3],
       [1, 2, 8, 7, 6],
       [2, 6, 9, 0, 3]])"""


#### seed
np.random.seed(666)
np.random.randint(0, 10, size=(2, 5))
""" array([[2, 6, 9, 4, 3],
       [1, 0, 8, 7, 5]]) """


### random
np.random.random((2,5))
""" array([[0.192892  , 0.70084475, 0.29322811, 0.77447945, 0.00510884],
       [0.11285765, 0.11095367, 0.24766823, 0.0232363 , 0.72732115]]) """


### random.normal
np.random.normal(10, 100)  # 均值：10, 方差：100

np.random.normal(0, 1, (3, 5)) ## 均值：0, 方差：1
""" 
array([[-0.57577075, -1.68290077,  0.22918525, -1.75662522,  0.84463262],
       [ 0.27721986,  0.85290153,  0.1945996 ,  1.31063772,  1.5438436 ],
       [-0.52904802, -0.6564723 , -0.2015057 , -0.70061583,  0.68713795]])  """


#### reshape
np.arange(10).reshape(2, 5)
""" array([[0, 1, 2, 3, 4],
       [5, 6, 7, 8, 9]])  """

np.arange(10).ndim    # 1

np.arange(10).reshape(2, 5).ndim    # 2

np.arange(10).shape     # (10,)

np.arange(10).reshape(2, 5).shape   # (2, 5)


#### 数据访问
x = np.arange(10)
X = np.arange(10).reshape(2, 5)

print(x[-1])
print(X[1, 3])

## 取最后一列
X = np.arange(15).reshape(3, 5)
"""
array([[ 0,  1,  2,  3,  4],
       [ 5,  6,  7,  8,  9],
       [10, 11, 12, 13, 14]]) """

## 取最后一列
X[:,:-1]


## 取第一行
X[0:1,:]
X[0, :]

## 取第一列
X[:,0:1]
X[:, 0]

## 转置
X[::-1,::-1]
"""
array([[14, 13, 12, 11, 10],
       [ 9,  8,  7,  6,  5],
       [ 4,  3,  2,  1,  0]])   """

##
x = array([0, 1, 2, 3, 4, 5, 6, 7, 8, 9])

x.reshape(1, 10)    # array([[0, 1, 2, 3, 4, 5, 6, 7, 8, 9]])
x.reshape(-1, 10)   # array([[0, 1, 2, 3, 4, 5, 6, 7, 8, 9]])
x.reshape(10, 1)
"""
array([[0],
       [1],
       [2],
       [3],
       [4],
       [5],
       [6],
       [7],
       [8],
       [9]])  """
x.reshape(10, -1)




""" [skl.datasets] """

from sklearn.datasets import load_iris
data = load_iris()
data.target[[10, 25, 50]]
# array([0, 0, 1])
list(data.target_names)
# ['setosa', 'versicolor', 'virginica']
data.feature_names
# ['sepal length (cm)', 'sepal width (cm)', 'petal length (cm)', 'petal width (cm)']
data.data[:1]
# array([[5.1, 3.5, 1.4, 0.2]])
data.filename
# 'C:\\ProgramData\\Anaconda3\\lib\\site-packages\\sklearn\\datasets\\data\\iris.csv'

""" [tqdm, 进度条] """
import time
from tqdm import tqdm, trange

trange(i)是tqdm(range(i))的一种简单写法
for i in trange(100):
    time.sleep(0.05)

for i in tqdm(range(100), desc='Processing'):
    time.sleep(0.05)
    
dic = ['a', 'b', 'c', 'd', 'e']
pbar = tqdm(dic)
for i in pbar:
    pbar.set_description('Processing '+i)
    time.sleep(0.2)



""" [ML基础知识总结] """

#### [数据划分] 
# https://www.datarobot.com/jp/blog/summary-of-ml-partitioning-part-2/
# 1. 留出法 Hold-out               ->  数据量大，一次训练即可
# 2. K折交叉验证 KFLOD CV          ->  数据量中规中矩（时间开销大） Kaggle用得较多
# 3. 自主采样 Booststrap           ->  Kaggle用的比较少


#### 树模型适合场景
## 结构化数据：1. 类别字段多  2. 聚合字段多
## XGBoost： 1. 较早的高阶树模型，精度较好 2.缺点：训练时间较长，对类别特征直出不友好 3.接口：sklearn接口、原生接口
## LightGBM(微软)： 1.对节点分裂改进后速度提升 2.缺点：加入了随机性，精度、稳定性有所下降 3.接口：sklearn接口、原生接口
## CatBoost(yandex)： 1.支持类别和字符串分裂 2.缺点：容易过拟合 3.接口：sklearn接口、原生接口




""" [XGBoost] """
from catboost.datasets import titanic
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import StratifiedKFold
import xgboost as xgb
import warnings
warnings.filterwarnings('ignore')
import pandas as pd


skf = StratifiedKFold(n_splits=3, shuffle=True, random_state=0)
train_titanic, test_titanic = titanic()
train_titanic.shape, test_titanic.shape

# 数据预处理
train_titanic.drop(['PassengerId', 'Name', 'Cabin'], axis=1, inplace=True)
test_titanic.drop(['PassengerId', 'Name', 'Cabin'], axis=1, inplace=True)

cat_features = list(train_titanic.select_dtypes(include=['object']).columns)

def data_proc(data, cat_features):
  for s in data.columns:
      if s in cat_features:
          data[s].fillna(value='no_mark', inplace=True)
      else:
          data[s].fillna(value=999, inplace=True)

data_proc(train_titanic, cat_features)
data_proc(test_titanic, cat_features)
cat_features = list(train_titanic.select_dtypes(include=['object']).columns)
cat_features



train_titanic['Ticket'] = [s.split(' ')[0] for s in train_titanic['Ticket']]
train_titanic['Ticket'] = ['is_digit' if s[0].isdigit() else s for s in train_titanic['Ticket']]

test_titanic['Ticket'] = [s.split(' ')[0] for s in test_titanic['Ticket']]
test_titanic['Ticket'] = ['is_digit' if s[0].isdigit() else s for s in test_titanic['Ticket']]


train_titanic['Sex'].unique()

def cat_factorize(df, cat_features):
    for cat_feature in cat_features:
        df[cat_feature] = pd.factorize(df[cat_feature], sort=True)[0]

cat_factorize(train_titanic,cat_features)
cat_factorize(test_titanic,cat_features)



X, y = train_titanic.drop(['Survived'], axis=1), train_titanic['Survived'].to_frame()
X_test = test_titanic.copy()

# train_test_split
from sklearn.model_selection import train_test_split
X_train, X_val, y_train, y_val = train_test_split(X, y, random_state=42, test_size=0.3, shuffle=True)
X_train.shape, X_val.shape, y_train.shape, y_val.shape



# %%time
# LightXGBの初期化
model = xgb.XGBClassifier()
#pprint.pprint(model.get_params())
# パラメーターを設定する
param_grid = {"max_depth": [ 3, 6, 10,25], #10, 25,
              "learning_rate" : [0.0001,0.001,0.01], # 0.05,0.1
              "min_child_weight" : [1,3,6],
              "n_estimators": [100,200,300], # 500
              "subsample": [0.5,0.75,0.9],
              "gamma":[0,0.1,0.2],
              "eta": [0.3,0.15,0.10]
             }
# パラメータチューニングをグリッドサーチで行うために設定する
## このGridSearchCV には注意が必要 scoring は そのスコアを基準にして最適化する
grid_result = GridSearchCV(estimator = model,
                           param_grid = param_grid,
                           scoring = 'balanced_accuracy',
                           cv = skf,
                           verbose=3,
                           return_train_score = True,
                           n_jobs = -1)
grid_result.fit(X_train, y_train)


grid_result.best_estimator_

grid_result.best_score_

grid_result.best_params_

grid_result.predict(test_titanic)[:3]


""" [CatBoost] """
from catboost.datasets import titanic
import warnings
warnings.filterwarnings('ignore')
import numpy as np
import pandas as pdb


train_titanic, test_titanic = titanic()
print(type(train_titanic)), train_titanic.shape, test_titanic.shape

list(train_titanic.columns)
train_titanic.head(2)
test_titanic.head(2)

train_titanic['Survived'].value_counts()
train_titanic.describe(include='all')

train_titanic.drop(['PassengerId', 'Name', 'Cabin'], axis=1, inplace=True)
train_titanic.head(2)

test_titanic.drop(['PassengerId', 'Name', 'Cabin'], axis=1, inplace=True)
test_titanic.head(2)

# 1. Data processing
cat_features = list(train_titanic.select_dtypes(include=['object']).columns)
cat_featur
#->  ['Sex', 'Ticket', 'Embarked']

def data_proc(data, cat_features):
  for s in data.columns:
      if s in cat_features:
          data[s].fillna(value='no_mark', inplace=True)
      else:
          data[s].fillna(value=999, inplace=True)

data_proc(train_titanic, cat_features)
data_proc(test_titanic, cat_features)

train_titanic.isna().sum()
# Survived    0
# Pclass      0
# Sex         0
# Age         0
# SibSp       0
# Parch       0
# Ticket      0
# Fare        0
# Embarked    0
# dtype: int64

train_titanic['Ticket'] = [s.split(' ')[0] for s in train_titanic['Ticket']]
train_titanic['Ticket'] = ['is_digit' if s[0].isdigit() else s for s in train_titanic['Ticket']]

test_titanic['Ticket'] = [s.split(' ')[0] for s in test_titanic['Ticket']]
test_titanic['Ticket'] = ['is_digit' if s[0].isdigit() else s for s in test_titanic['Ticket']]

def cat_factorize(df, cat_features):
    for cat_feature in cat_features:
        df[cat_feature] = pd.factorize(df[cat_feature], sort=True)[0]

cat_factorize(train_titanic,cat_features)
cat_factorize(test_titanic,cat_features)

X, y = train_titanic.drop(['Survived'], axis=1), train_titanic['Survived'].to_frame()
X_test = test_titanic.copy()
X.shape, y.shape, X_test.shape


# 2. Grid search
from catboost import CatBoostClassifier
warnings.filterwarnings('ignore')
import numpy as np
from sklearn.model_selection import GridSearchCV

params = {
    'depth': [3, 4, 5, 6, 7],
    'learning_rate': [0.03, 0.1, 0.15],
    'l2_leaf_reg': [2, 3, 4, 5, 6, 7],
    'iterations': [300],
#     'task_type':['GPU'],
#     'loss_function':['MultiClass'],
}

ctb = CatBoostClassifier(eval_metric='AUC', logging_level='Silent', cat_features=cat_features)
# ctb_grid_search = GridSearchCV(ctb, param_grid=params, scoring='roc_auc',  cv=3, verbose=False)
ctb_grid = GridSearchCV(ctb, param_grid=params, scoring='accuracy',  cv=5, verbose=False)

ctb_grid.fit(X, y, eval_set=(X, y))
# ctb_grid_search.grid_scores_,  ctb_grid_search.best_params_,  ctb_grid_search.best_score_
ctb_grid.best_score_, ctb_grid.best_estimator_.get_params()

ctb_grid.predict(X_test)[:10]
# ->  array([0, 1, 0, 0, 1, 0, 1, 0, 1, 0], dtype=int64)
ctb_grid.predict_proba(X_test)[:3]


# StratifiedKFlod
from sklearn.model_selection import StratifiedKFold
skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=0)

from catboost import CatBoostClassifier
warnings.filterwarnings('ignore')
import numpy as np
from sklearn.model_selection import GridSearchCV

params = {
    'depth': [3, 4, 5, 6, 7],
    'learning_rate': [0.03, 0.1, 0.15],
    'l2_leaf_reg': [2, 3, 4, 5, 6, 7],
    'iterations': [300],
#     'task_type':['GPU'],
#     'loss_function':['MultiClass'],
}

ctb = CatBoostClassifier(eval_metric='AUC', logging_level='Silent', cat_features=cat_features)
# ctb_grid_search = GridSearchCV(ctb, param_grid=params, scoring='roc_auc',  cv=3, verbose=False)
ctb_grid = GridSearchCV(ctb, param_grid=params, scoring='accuracy',  cv=skf, verbose=False)

ctb_grid.fit(X, y, eval_set=(X, y))
# ctb_grid_search.grid_scores_,  ctb_grid_search.best_params_,  ctb_grid_search.best_score_
ctb_grid.best_score_, ctb_grid.best_estimator_.get_params()


""" [DL-Basic-boston] """
import numpy as np # linear algebra
import pandas as pd # data processing, CSV file I/O (e.g. pd.read_csv)
import warnings
warnings.filterwarnings('ignore')
import torch

from sklearn.datasets import load_boston
data = load_boston()
X = data.data
Y = data.target
X.shape, Y.shape
# -> ((506, 13), (506,))

from sklearn.model_selection import train_test_split
X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=10, random_state=42, shuffle=True)
X_train.shape, X_test.shape, Y_train.shape, Y_test.shape
# -> ((496, 13), (10, 13), (496,), (10,))


#### 单层结构
# net
class Net(torch.nn.Module):
    def __init__(self, n_feature, n_output):
        super(Net, self).__init__()
        self.predict = torch.nn.Linear(n_feature, n_output)
    
    def forward(self, x):
        out = self.predict(x)
        return out
    
net = Net(13, 1) #13->特征数量, 1->一个输出值        

# loss
loss_func = torch.nn.MSELoss()   # MSELoss->均方误差损失

# optimizer
# optimizer = torch.optim.SGD(net.parameters(), lr=0.0001) # lr -> 学习率
optimizer = torch.optim.Adam(net.parameters(), lr=0.0001) # lr -> 学习率

# training
for i in range(10000):
    x_data = torch.tensor(X_train, dtype=torch.float32)
    y_data = torch.tensor(Y_train, dtype=torch.float32)
    
    # 前向运算
    pred = net.forward(x_data)
    pred = torch.squeeze(pred)
    loss = loss_func(pred, y_data) * 0.001
    optimizer.zero_grad()
    loss.backward()
    optimizer.step()
    if i%4999==0:
        print(r'ite:{}, loss:{}'.format(i, loss))
        print(pred[:10])
        print(y_data[:10])

# ite:9998, loss:0.07451146841049194
# tensor([26.0727, 24.7244,  9.1978, 20.5771, 23.9176, 19.6255, 30.0153, 21.2493,
#         32.4744, 22.0688], grad_fn=<SliceBackward0>)
# tensor([21.5000, 18.9000,  7.0000, 21.2000, 18.5000, 29.8000, 18.8000, 10.2000,
#         50.0000, 14.1000])


#### 加隐藏层
# net
class Net(torch.nn.Module):
    def __init__(self, n_feature, n_output):
        super(Net, self).__init__()
        self.hidden = torch.nn.Linear(n_feature, 100)
        self.predict = torch.nn.Linear(100, n_output)
    
    def forward(self, x):
        out = self.hidden(x)
        out = torch.relu(out)
        out = self.predict(out)
        return out
    
net = Net(13, 1) #13->特征数量, 1->一个输出值        

# loss
loss_func = torch.nn.MSELoss()   # MSELoss->均方误差损失

# optimizer
# optimizer = torch.optim.SGD(net.parameters(), lr=0.0001) # lr -> 学习率
optimizer = torch.optim.Adam(net.parameters(), lr=0.0001) # lr -> 学习率

# training
for i in range(10000):
    x_data = torch.tensor(X_train, dtype=torch.float32)
    y_data = torch.tensor(Y_train, dtype=torch.float32)
    
    # 前向运算
    pred = net.forward(x_data)
    pred = torch.squeeze(pred)
    loss = loss_func(pred, y_data) * 0.001
    optimizer.zero_grad()
    loss.backward()
    optimizer.step()
    if i%4999==0:
        print(r'ite:{}, loss_train:{}'.format(i, loss))
        print(pred[:10])
        print(y_data[:10])        
 

    # test    
    x_data = torch.tensor(X_test, dtype=torch.float32)
    y_data = torch.tensor(Y_test, dtype=torch.float32)
    
    # 前向运算
    pred = net.forward(x_data)
    pred = torch.squeeze(pred)
    loss = loss_func(pred, y_data) * 0.001
    optimizer.zero_grad()
    loss.backward()
    if i%4999==0:
        print(r'ite:{}, loss_test:{}'.format(i, loss))
        print(pred[:10])
        print(y_data[:10])   

# ite:9998, loss_test:0.005073814187198877
# tensor([24.7437, 35.0739, 17.0265, 24.1485, 15.3470, 21.1183, 18.4505, 16.1415,
#         19.5541, 21.4789], grad_fn=<SliceBackward0>)
# tensor([23.6000, 32.4000, 13.6000, 22.8000, 16.1000, 20.0000, 17.8000, 14.0000,
#         19.6000, 16.8000])


#### 调用模型
# 保存、调用模型（模型方式）
torch.save(net, 'model.pkl')
torch.load("model.pkl")

# 保存、调用模型（模型参数方式）
torch.save(net.state_dict(), "params.pkl")
net.load_state_dict("params.pkl")



torch.load("model.pkl")




""" [pytorch基础] """
#### pip install torch===1.7.1 torchvision===0.4.1 -i https://pypi.tuna.tsinghua.edu.cn/simple
#### csv转tensor  https://copyfuture.com/blogs-details/20211205151623638m
#### https://www.kaggle.com/code/tianyc1026/dl-minst/edit/run/110039162
dt = pd.read_csv(r'/kaggle/input/digit-recognizer/train.csv')
dt.head(2)
## DataFrame -> Numpy
dt.values
X, Y = dt.values[:,1:-1].astype(float), dt.values[:,:1].astype(int)
print(X.shape, Y.shape)
# -> (42000, 783) (42000, 1)

X, Y = torch.FloatTensor(X), torch.LongTensor(Y)
from sklearn.model_selection import train_test_split
X_train, X_test, Y_train, Y_test = train_test_split(X, Y, shuffle=True, test_size=0.2, random_state=42)

train_dataset = torch.utils.data.TensorDataset(X_train, Y_train)
test_dataset = torch.utils.data.TensorDataset(X_test, Y_test)


#### torch的dataset
import torch
import torchvision.datasets as dataset
import torchvision.transforms as transforms
import torch.utils.data as data_utils

# Data
train_data = dataset.MNIST(root=PATH,train=True,\
                          transform=transforms.ToTensor(),\
                          download=False)
train_data

test_data = dataset.MNIST(root=PATH,train=False,\
                          transform=transforms.ToTensor(),\
                          download=False)
test_data
# Dataset MNIST
#     Number of datapoints: 10000
#     Root location: C:\Users\tianyunchuan\iCloudDrive\_data_\data_class\DL_PyTorch_class_163_Morvan\mnist
#     Split: Test
#     StandardTransform
# Transform: ToTensor()

# batch size
train_loader = data_utils.DataLoader(dataset=train_data,\
                                    batch_size=64,\
                                    shuffle=True)
test_loader = data_utils.DataLoader(dataset=test_data,\
                                    batch_size=64,\
                                    shuffle=True)


#### batchsize
"""
    https://blog.csdn.net/weixin_44211968/article/details/123739994
    批训练，把数据变成一小批一小批数据进行训练。
    DataLoader就是用来包装所使用的数据，每次抛出一批数据
"""
import torch
import torch.utils.data as Data

BATCH_SIZE = 4       # 批训练的数据个数

x = torch.linspace(1, 10, 10)   # 训练数据
print(x)
y = torch.linspace(10, 1, 10)   # 标签
print(y)
# -> torch.utils.data.dataset.TensorDataset

#####!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
##X, Y = torch.FloatTensor(X), torch.LongTensor(Y)


# 把数据放在数据库中
torch_dataset = Data.TensorDataset(x, y)  # 对给定的 tensor 数据，将他们包装成 dataset

loader = Data.DataLoader(
    # 从数据库中每次抽出batch size个样本
    dataset=torch_dataset,       # torch TensorDataset format
    batch_size=BATCH_SIZE,       # mini batch size
    shuffle=True,                # 要不要打乱数据 (打乱比较好)
    num_workers=2,               # 多线程来读数据
)

def show_batch():
    for epoch in range(3):
        for step, (batch_x, batch_y) in enumerate(loader):
            # training
            print("steop:{}, batch_x:{}, batch_y:{}".format(step, batch_x, batch_y))

show_batch()

# step:0, batch_x:tensor([ 1.,  8., 10.,  6.]), batch_y:tensor([10.,  3.,  1.,  5.])
# step:1, batch_x:tensor([7., 9., 5., 4.]), batch_y:tensor([4., 2., 6., 7.])
# step:2, batch_x:tensor([3., 2.]), batch_y:tensor([8., 9.])

# step:0, batch_x:tensor([2., 5., 9., 8.]), batch_y:tensor([9., 6., 2., 3.])
# step:1, batch_x:tensor([1., 3., 7., 6.]), batch_y:tensor([10.,  8.,  4.,  5.])
# step:2, batch_x:tensor([10.,  4.]), batch_y:tensor([1., 7.])

# step:0, batch_x:tensor([8., 7., 5., 2.]), batch_y:tensor([3., 4., 6., 9.])
# step:1, batch_x:tensor([ 9.,  3.,  4., 10.]), batch_y:tensor([2., 8., 7., 1.])
# step:2, batch_x:tensor([1., 6.]), batch_y:tensor([10.,  5.])


#### torchvision报错
# 解决方法

# 方法一（推荐）
# torchvision在运行时需要PILLOW_VERSION函数，而新版的pillow把PILLOW_VERSION函数换成了__version__函数

# 那么就把torchvision代码文件中的from PIL import Image, ImageOps, ImageEnhance,PILLOW_VERSION改为from PIL import Image, ImageOps, ImageEnhance, __version__

# 操作方法如下：

    # 按照提示路径C:\Users\mayuhuaw\software\Anaconda3_2020_11\anaconda\envs\pytorch\lib\site-packages\torchvision\transforms\functional.py
    # 打开 C:\Users\mayuhuaw\software\Anaconda3_2020_11\anaconda\envs\pytorch\lib\site-packages\torchvision\transforms文件夹中的functional.py文件

#### cv2
# anaconda prompt 管理员打开
# pip install opencv-contrib-python -i http://pypi.douban.com/simple --trusted-host pypi.douban.com
# pip install opencv-python -i http://pypi.douban.com/simple/ --trusted-host pypi.douban.com
# opencv-python 下载安装
# pip install -U numpy==1.12.0



""" [DL] """
#### 卷积模型
## 卷积层 -> 池化层(Pooling) -> 激活层-BN层-FC层-损失层
## 激活层：增加网络的非线性

#### 经典卷积模型
## 简单神经网络
# LeNet
# VGG
# AlexNet
## 复杂神经网络
# ResNet (何凯明)
# InceptionNet V1-V4 (用的比较少)
# DenseNet
## 轻量型神经网络 (重要！)
# MobileNet V1-V3  (Google团队开发)
# ShuffleNet
# SqueezeNet
## 多分支网络结构
# SiameseNet
# TripletNet
## attention网络结构 （用于循环网络比较多)

#### 学习率
# 指数衰减

#### 优化器
# GD、BGD、SGD、MBGD (引入了随机性和噪声)
# Momentum、NAG等 (加入动量原则、具有加速梯度下降的作用)
# AdaGrad、RMSProp、Adam、Adadelta （自适应学习)
# torch.optim.Adam
# 优化器推荐使用 torch.optim.Adam, lr=0.001 (学习率初始值), 学习率指数衰减: torch.optim.lr_scheduler.ExponentialLR 

